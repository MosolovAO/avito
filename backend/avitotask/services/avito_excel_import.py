from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from avitotask.models import AvitoListing, ProductOptions

from django.db import transaction
from django.utils import timezone

HEADER_ROW = 2
REQUIRED_ROW = 3
FORMAT_ROW = 4
DATA_START_ROW = 5

REFERENCE_SHEET_PREFIX = "Спр-"
INSTRUCTION_SHEET_NAME = "Инструкция"

IMAGE_URLS_SEPARATOR = " | "

MODEL_FIELD_KEYS = {
    "row_id",
    "avito_id",
    "title",
    "description",
    "image_urls",
    "address",
    "status",
}

BASE_DATA_KEYS = {
    "ListingFee",
    "ManagerName",
    "ContactPhone",
    "ContactMethod",
    "Category",
    "Price",
    "GoodsType",
    "AdType",
    "Condition",
    "Availability",
    "EMail",
    "CompanyName",
    "DateEnd",
}

CORE_COLUMN_MAP = {
    "Уникальный идентификатор объявления": "row_id",
    "Номер объявления на Авито": "avito_id",
    "Название объявления": "title",
    "Описание объявления": "description",
    "Ссылки на фото": "image_urls",
    "Адрес": "address",
    "AvitoStatus": "status",

    # Базовые поля автозагрузки. Их лучше маппить стабильно даже без ProductOptions.
    "Способ размещения": "ListingFee",
    "Контактное лицо": "ManagerName",
    "Номер телефона": "ContactPhone",
    "Способ связи": "ContactMethod",
    "Категория": "Category",
    "Цена": "Price",
    "Вид товара": "GoodsType",
    "Вид объявления": "AdType",
    "Состояние": "Condition",
    "Доступность": "Availability",
    "Почта": "EMail",
    "Название компании": "CompanyName",
    "AvitoDateEnd": "DateEnd",
}


@dataclass(frozen=True)
class AvitoExcelColumn:
    ru_title: str
    technical_title: str | None
    required: str
    value_format: str


@dataclass(frozen=True)
class AvitoExcelImportResult:
    total_rows: int
    skipped_rows: int
    created_listings: int
    updated_listings: int
    rows_with_errors: int
    unmapped_columns: list[str]


@dataclass(frozen=True)
class AvitoExcelRowPreview:
    sheet_name: str
    category_path: str
    row_number: int
    row_id: str
    avito_id: str
    title: str
    status: str
    raw_data: dict[str, Any]
    mapped_data: dict[str, Any]
    unmapped_data: dict[str, Any]
    column_data: list[dict[str, Any]]
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class AvitoExcelPreviewResult:
    total_sheets: int
    total_rows: int
    rows_with_errors: int
    categories: list[str]
    unmapped_columns: list[str]
    rows: list[AvitoExcelRowPreview]


class AvitoExcelImportError(Exception):
    pass


def preview_avito_excel_file(file_obj) -> AvitoExcelPreviewResult:
    """
    Читает XLSX из кабинета Avito и возвращает preview без записи в БД.

    Важно:
    - русские заголовки маппятся через ProductOptions.option_title_ru;
    - внутренние ключи берутся из ProductOptions.option_title_en;
    - неизвестные колонки не теряются, а попадают в unmapped_data.
    """

    workbook = load_workbook(file_obj, read_only=False, data_only=True)
    option_map = get_product_option_map()

    rows: list[AvitoExcelRowPreview] = []
    categories: list[str] = []
    unmapped_columns: list[str] = set()
    total_sheets = 0

    for worksheet in workbook.worksheets:
        if should_skip_sheet(worksheet.title):
            continue

        category_path = normalize_cell_value(worksheet.cell(row=1, column=1).value)
        headers = read_row_values(worksheet, HEADER_ROW)

        if not category_path or not has_real_headers(headers):
            continue

        total_sheets += 1
        categories.append(category_path)

        required_flags = read_row_values(worksheet, REQUIRED_ROW)
        format_flags = read_row_values(worksheet, FORMAT_ROW)
        columns = build_columns(
            headers=headers,
            required_flags=required_flags,
            format_flags=format_flags,
            option_map=option_map,
        )

        for row_number in range(DATA_START_ROW, worksheet.max_row + 1):
            raw_data = read_listing_row(worksheet, row_number, columns)

            if not has_real_row_data(raw_data):
                continue

            row_preview = build_row_preview(
                sheet_name=worksheet.title,
                category_path=category_path,
                row_number=row_number,
                raw_data=raw_data,
                columns=columns
            )
            rows.append(row_preview)
            unmapped_columns.update(row_preview.unmapped_data.keys())

    return AvitoExcelPreviewResult(
        total_sheets=total_sheets,
        total_rows=len(rows),
        rows_with_errors=sum(1 for row in rows if row.errors),
        categories=sorted(set(categories)),
        unmapped_columns=sorted(unmapped_columns),
        rows=rows,
    )


def get_product_option_map() -> dict[str, str]:
    return {
        option.option_title_ru.strip(): option.option_title_en.strip()
        for option in ProductOptions.objects.exclude(option_title_en="")
        if option.option_title_ru and option.option_title_en
    }


def split_mapped_data(mapped_data: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    model_data = {}
    base_data = {}
    option_data = {}

    for key, value in mapped_data.items():
        if key in MODEL_FIELD_KEYS:
            model_data[key] = value
        elif key in BASE_DATA_KEYS:
            base_data[key] = value
        else:
            option_data[key] = value

    return model_data, base_data, option_data


def should_skip_sheet(sheet_name: str) -> bool:
    return (
            sheet_name == INSTRUCTION_SHEET_NAME
            or sheet_name.startswith(REFERENCE_SHEET_PREFIX)
    )


def read_row_values(worksheet, row_number: int) -> list[str]:
    return [
        normalize_cell_value(worksheet.cell(row=row_number, column=column).value)
        for column in range(1, worksheet.max_column + 1)
    ]


def has_real_headers(headers: list[str]) -> bool:
    return any(value for value in headers)


def has_real_row_data(raw_data: dict[str, Any]) -> bool:
    return any(
        normalize_cell_value(value) != ""
        for value in raw_data.values()
    )


def build_columns(
        *,
        headers: list[str],
        required_flags: list[str],
        format_flags: list[str],
        option_map: dict[str, str],
) -> list[AvitoExcelColumn]:
    columns = []

    for index, ru_title in enumerate(headers):
        if not ru_title:
            continue

        technical_title = CORE_COLUMN_MAP.get(ru_title) or option_map.get(ru_title)

        columns.append(
            AvitoExcelColumn(
                ru_title=ru_title,
                technical_title=technical_title,
                required=value_at(required_flags, index),
                value_format=value_at(format_flags, index),
            )
        )

    return columns


def read_listing_row(worksheet, row_number: int, columns: list[AvitoExcelColumn]) -> dict[str, Any]:
    result = {}

    for column_index, column in enumerate(columns, start=1):
        value = worksheet.cell(row=row_number, column=column_index).value
        normzalized_value = normalize_cell_value(value)

        if normzalized_value != "":
            result[column.ru_title] = normzalized_value

    return result


def build_row_preview(
        *,
        sheet_name: str,
        category_path: str,
        row_number: int,
        raw_data: dict[str, Any],
        columns: list[AvitoExcelColumn],
) -> AvitoExcelRowPreview:
    mapped_data = {}
    unmapped_data = {}

    column_by_ru_title = {
        column.ru_title: column
        for column in columns
    }

    for ru_title, value in raw_data.items():
        column = column_by_ru_title.get(ru_title)

        if column and column.technical_title:
            mapped_data[column.technical_title] = normalize_field_value(
                technical_title=column.technical_title,
                value=value,
            )
            continue

        if column and is_required_column(column.required):
            unmapped_data[ru_title] = value

    row_id = str(mapped_data.get("row_id") or "")
    avito_id = str(mapped_data.get("avito_id") or "")
    title = str(mapped_data.get("title") or "")
    status = str(mapped_data.get("status") or "")

    errors = validate_preview_row(
        row_id=row_id,
        avito_id=avito_id,
        title=title,
        row_number=row_number,
        sheet_name=sheet_name
    )

    return AvitoExcelRowPreview(
        sheet_name=sheet_name,
        category_path=category_path,
        row_number=row_number,
        row_id=row_id,
        avito_id=avito_id,
        title=title,
        status=status,
        raw_data=raw_data,
        mapped_data=mapped_data,
        unmapped_data=unmapped_data,
        errors=errors,
        column_data=[
            {
                "ru_title": column.ru_title,
                "technical_title": column.technical_title,
                "required": column.required,
                "value_format": column.value_format,
            }
            for column in columns
        ],
    )


def validate_preview_row(
        *,
        row_id: str,
        avito_id: str,
        title: str,
        row_number: int,
        sheet_name: str,
) -> list[str]:
    errors = []

    if not row_id:
        errors.append(f"{sheet_name}:{row_number}: пустой уникальный идентификатор объявления.")

    if not avito_id:
        errors.append(f"{sheet_name}:{row_number}: пустой номер объявления на Авито.")

    if not title:
        errors.append(f"{sheet_name}:{row_number}: пустое название объявления.")

    return errors


def normalize_field_value(*, technical_title: str, value: Any) -> Any:
    if technical_title == "image_urls":
        return split_image_urls(value)

    return value


def split_image_urls(value: Any) -> list[str]:
    if not value:
        return []

    return [
        item.strip()
        for item in str(value).split(IMAGE_URLS_SEPARATOR)
        if item.strip()
    ]


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def value_at(values: list[str], index: int) -> str:
    if index >= len(values):
        return ""

    return values[index]


def is_required_column(required_value: str) -> bool:
    normalized_value = normalize_cell_value(required_value).lower()

    return normalized_value in {
        "да",
        "yes",
        "true",
        "1",
        "обязательно",
        "обязательный",
        "required",
    }


@transaction.atomic
def import_avito_excel_file(*, workspace, avito_account, file_obj) -> AvitoExcelImportResult:
    if avito_account.workspace_id != workspace.id:
        raise AvitoExcelImportError("Аккаунт Avito принадлежит другому workspace.")

    preview = preview_avito_excel_file(file_obj)

    created_listings = 0
    updated_listings = 0
    skipped_rows = 0

    for row in preview.rows:
        if row.errors:
            skipped_rows += 1
            continue

        model_data, base_data, option_data = split_mapped_data(row.mapped_data)

        avito_id = model_data.get("avito_id")
        if not avito_id:
            skipped_rows += 1
            continue

        listing, was_created = AvitoListing.objects.update_or_create(
            workspace=workspace,
            avito_account=avito_account,
            avito_id=str(avito_id),
            defaults={
                "source": AvitoListing.Source.AVITO_EXCEL,
                "management_status": AvitoListing.ManagementStatus.MANAGED,
                "desired_status": AvitoListing.DesiredStatus.PUBLISH,
                "publication": None,
                "row_id": model_data.get("row_id") or "",
                "status": model_data.get("status") or "",
                "title": model_data.get("title") or "",
                "description": model_data.get("description") or "",
                "address": model_data.get("address") or "",
                "image_urls": model_data.get("image_urls") or [],
                "base_data": base_data,
                "option_data": option_data,
                "raw_data": row.raw_data,
                "unmapped_data": row.unmapped_data,
                "sheet_name": row.sheet_name,
                "category_path": row.category_path,
                "imported_payload": {
                    "source": AvitoListing.Source.AVITO_EXCEL,
                    "sheet_name": row.sheet_name,
                    "category_path": row.category_path,
                    "row_number": row.row_number,
                    "raw_data": row.raw_data,
                    "mapped_data": row.mapped_data,
                    "unmapped_data": row.unmapped_data,
                    "column_data": row.column_data,
                },
                "last_seen_at": timezone.now(),
            },
        )

        if was_created:
            created_listings += 1
        else:
            updated_listings += 1

    return AvitoExcelImportResult(
        total_rows=preview.total_rows,
        skipped_rows=skipped_rows,
        created_listings=created_listings,
        updated_listings=updated_listings,
        rows_with_errors=preview.rows_with_errors,
        unmapped_columns=preview.unmapped_columns,
    )
