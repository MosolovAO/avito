from django.test import TestCase
from avitotask.services.ad_editing import update_ad_creative, update_ad_publication

from unittest.mock import call, patch
from django.utils import timezone

from rest_framework.test import APIClient

from io import BytesIO
from openpyxl import Workbook

from avitotask.services.avito_import import import_avito_listings_for_account, upsert_avito_listing
from avitotask.services.avito_listing_editing import (
    bulk_update_avito_listing_desired_status,
    bulk_update_avito_listing_management_status,
)
from avitotask.services.avito_listing_lifecycle import build_avito_listing_lifecycle_report

from avitotask.services.avito_autoload import link_publications_to_avito_ids_for_account

from django.urls import reverse
from system.celery import app as celery_app

from datetime import datetime, date, timedelta

from avitotask.services.ad_schedule import (
    AdScheduleError,
    calculate_next_run_at,
    normalize_schedule,
    run_due_ad_generation_tasks,
)
from avitotask.services.avito_autoload_report_fetch import (
    sync_last_completed_autoload_report_for_account,
)
from avitotask.services.ad_cleanup import archive_stale_publications
from avitotask.services.avito_stats import import_avito_listing_daily_stats_for_account
from avitotask.services.ad_export import (
    build_publication_export_row,
    export_avito_account_publications_to_csv,
)
from avitotask.tasks import export_avito_account_csv_task, export_dirty_avito_accounts_csv_task, \
    import_avito_account_listings_task, link_avito_account_publications_task, import_avito_account_daily_stats_task, \
    sync_last_completed_avito_autoload_report_task

from accounts.models import User, Workspace, WorkspaceMembership
from avitotask.models import (
    AdBatch,
    AdCreative,
    AvitoListing,
    AvitoListingDailyStats,
    AdGenerationTask,
    AdPublication,
    AvitoAccount,
    AvitoOAuthToken,
    ProductOptions,
    AdGenerationTaskRun,
    AdImageAsset,
)
from avitotask.services.ad_generation import (
    create_manual_mass_posting,
    generate_ads_from_task, AdGenerationError, build_creative_dedupe_data,
)
from avitotask.services.ad_task_runner import run_autogeneration_task
from avitotask.services.avito_excel_import import (
    import_avito_excel_file,
    preview_avito_excel_file,
)
from avitotask.services.avito_autoload_report_sync import sync_avito_autoload_report
from avitotask.services.avito_api import connect_avito_account_from_token, extract_avito_user_id

import csv
import tempfile
from pathlib import Path

from zoneinfo import ZoneInfo


class AvitoExcelImportFlowTests(TestCase):

    def test_last_completed_autoload_report_sync_links_publication_to_avito_listing(self):
        publication = self.create_publication_for_autoload_report(
            row_id="LAST-REPORT-ROW-001",
        )

        AvitoOAuthToken.objects.create(
            workspace=self.workspace,
            avito_account=self.avito_account,
            access_token="autoload-report-access-token",
            refresh_token="autoload-report-refresh-token",
            scope="autoload:reports",
        )

        class FakeResponse:
            status_code = 200
            text = "json"

            def __init__(self, payload):
                self.payload = payload

            def json(self):
                return self.payload

        class FakeSession:
            def __init__(self):
                self.calls = []

            def request(self, method, url, **kwargs):
                self.calls.append((method, url, kwargs))

                if url.endswith("/autoload/v3/reports/last_completed_report"):
                    return FakeResponse({
                        "id": "REPORT-001",
                        "status": "completed",
                    })

                if url.endswith("/autoload/v2/reports/REPORT-001/items"):
                    return FakeResponse({
                        "items": [
                            {
                                "ad_id": publication.row_id,
                                "avito_id": "7777777777",
                                "status": "accepted",
                            },
                        ],
                    })

                return FakeResponse({})

        session = FakeSession()

        result = sync_last_completed_autoload_report_for_account(
            avito_account=self.avito_account,
            session=session,
        )

        self.assertEqual(result.report_id, "REPORT-001")
        self.assertEqual(result.total_items_received, 1)
        self.assertEqual(result.sync_result.total_rows, 1)
        self.assertEqual(result.sync_result.accepted_rows, 1)
        self.assertEqual(result.sync_result.linked_publications, 1)
        self.assertEqual(result.sync_result.created_listings, 1)

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            publication=publication,
        )

        self.assertEqual(listing.avito_id, "7777777777")
        self.assertEqual(listing.row_id, publication.row_id)
        self.assertEqual(listing.source, AvitoListing.Source.SERVICE)

        self.assertEqual(session.calls[0][0], "GET")
        self.assertTrue(
            session.calls[0][1].endswith("/autoload/v3/reports/last_completed_report")
        )
        self.assertTrue(
            session.calls[1][1].endswith("/autoload/v2/reports/REPORT-001/items")
        )

    def test_last_completed_autoload_report_sync_is_idempotent(self):
        publication = self.create_publication_for_autoload_report(
            row_id="LAST-REPORT-IDEMPOTENT-001",
        )

        AvitoOAuthToken.objects.create(
            workspace=self.workspace,
            avito_account=self.avito_account,
            access_token="autoload-report-idempotent-token",
            refresh_token="autoload-report-idempotent-refresh",
            scope="autoload:reports",
        )

        class FakeResponse:
            status_code = 200
            text = "json"

            def __init__(self, payload):
                self.payload = payload

            def json(self):
                return self.payload

        class FakeSession:
            def request(self, method, url, **kwargs):
                if url.endswith("/autoload/v3/reports/last_completed_report"):
                    return FakeResponse({
                        "id": "REPORT-IDEMPOTENT-001",
                        "status": "completed",
                    })

                if url.endswith("/autoload/v2/reports/REPORT-IDEMPOTENT-001/items"):
                    return FakeResponse({
                        "items": [
                            {
                                "ad_id": publication.row_id,
                                "avito_id": "7777777788",
                                "status": "accepted",
                            },
                        ],
                    })

                return FakeResponse({})

        session = FakeSession()

        first_result = sync_last_completed_autoload_report_for_account(
            avito_account=self.avito_account,
            session=session,
        )
        second_result = sync_last_completed_autoload_report_for_account(
            avito_account=self.avito_account,
            session=session,
        )

        self.assertEqual(first_result.sync_result.created_listings, 1)
        self.assertEqual(second_result.sync_result.created_listings, 0)
        self.assertEqual(second_result.sync_result.updated_listings, 1)

        self.assertEqual(
            AvitoListing.objects.filter(
                workspace=self.workspace,
                avito_account=self.avito_account,
                publication=publication,
                avito_id="7777777788",
            ).count(),
            1,
        )

    def test_last_completed_autoload_report_task_updates_sync_state(self):
        publication = self.create_publication_for_autoload_report(
            row_id="LAST-REPORT-TASK-001",
        )

        AvitoOAuthToken.objects.create(
            workspace=self.workspace,
            avito_account=self.avito_account,
            access_token="autoload-report-task-token",
            refresh_token="autoload-report-task-refresh",
            scope="autoload:reports",
        )

        class FakeResponse:
            status_code = 200
            text = "json"

            def __init__(self, payload):
                self.payload = payload

            def json(self):
                return self.payload

        class FakeSession:
            def request(self, method, url, **kwargs):
                if url.endswith("/autoload/v3/reports/last_completed_report"):
                    return FakeResponse({
                        "id": "REPORT-TASK-001",
                        "status": "completed",
                    })

                if url.endswith("/autoload/v2/reports/REPORT-TASK-001/items"):
                    return FakeResponse({
                        "items": [
                            {
                                "ad_id": publication.row_id,
                                "avito_id": "7777777799",
                                "status": "accepted",
                            },
                        ],
                    })

                return FakeResponse({})

        result = sync_last_completed_avito_autoload_report_task(
            self.avito_account.id,
            session=FakeSession(),
        )

        self.avito_account.refresh_from_db()

        self.assertEqual(result["status"], "completed")
        self.assertEqual(result["report_id"], "REPORT-TASK-001")
        self.assertEqual(result["linked_publications"], 1)
        self.assertEqual(result["created_listings"], 1)

        self.assertEqual(self.avito_account.sync_status, AvitoAccount.SyncStatus.IDLE)
        self.assertIsNone(self.avito_account.sync_error)
        self.assertIsNotNone(self.avito_account.last_synced_at)
        self.assertEqual(self.avito_account.last_sync_total_received, 1)
        self.assertEqual(self.avito_account.last_sync_created_listings, 1)

        listing = AvitoListing.objects.get(publication=publication)
        self.assertEqual(listing.avito_id, "7777777799")

    def test_extract_avito_user_id_supports_root_id(self):
        self.assertEqual(
            extract_avito_user_id({"id": 94235311}),
            "94235311",
        )

    def test_extract_avito_user_id_supports_nested_result_id(self):
        self.assertEqual(
            extract_avito_user_id({"result": {"id": 94235311}}),
            "94235311",
        )

    def test_extract_avito_user_id_returns_empty_string_when_missing(self):
        self.assertEqual(
            extract_avito_user_id({"name": "Петр"}),
            "",
        )

    def test_ads_api_returns_listings_and_unlinked_publications_without_duplicates(self):
        linked_publication = self.create_publication_for_autoload_report()

        sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=[
                {
                    "Id": linked_publication.row_id,
                    "AvitoId": "9999999999",
                    "status": "accepted",
                }
            ],
        )

        unlinked_publication = self.create_publication_for_autoload_report(
            row_id="SERVICE-ROW-002",
        )

        client = APIClient()
        client.force_authenticate(user=self.user)
        client.defaults["HTTP_X_WORKSPACE_ID"] = str(self.workspace.id)

        url = reverse(
            "avito-account-ads-list",
            kwargs={"avito_account_id": self.avito_account.id},
        )

        response = client.get(
            url,
            {
                "page": 1,
                "page_size": 20,
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)

        data = response.json()
        results = data["results"]

        linked_publication_rows = [
            item
            for item in results
            if item["publication"] == linked_publication.id
        ]
        unlinked_publication_rows = [
            item
            for item in results
            if item["publication"] == unlinked_publication.id
        ]

        self.assertEqual(len(linked_publication_rows), 1)
        self.assertEqual(linked_publication_rows[0]["entity_type"], "avito_listing")
        self.assertEqual(linked_publication_rows[0]["avito_id"], "9999999999")

        self.assertEqual(len(unlinked_publication_rows), 1)
        self.assertEqual(unlinked_publication_rows[0]["entity_type"], "ad_publication")
        self.assertIsNone(unlinked_publication_rows[0]["avito_id"])

    def test_ads_api_filters_by_ad_publication_entity_type(self):
        linked_publication = self.create_publication_for_autoload_report()

        sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=[
                {
                    "Id": linked_publication.row_id,
                    "AvitoId": "9999999999",
                    "status": "accepted",
                }
            ],
        )

        unlinked_publication = self.create_publication_for_autoload_report(
            row_id="SERVICE-ROW-002",
        )

        client = APIClient()
        client.force_authenticate(user=self.user)
        client.defaults["HTTP_X_WORKSPACE_ID"] = str(self.workspace.id)

        url = reverse(
            "avito-account-ads-list",
            kwargs={"avito_account_id": self.avito_account.id},
        )

        response = client.get(
            url,
            {
                "entity_type": "ad_publication",
                "page": 1,
                "page_size": 20,
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)

        results = response.json()["results"]

        self.assertTrue(results)
        self.assertTrue(
            all(item["entity_type"] == "ad_publication" for item in results)
        )
        self.assertEqual(
            [item["publication"] for item in results],
            [unlinked_publication.id],
        )

    def test_ads_api_filters_items_without_avito_id(self):
        linked_publication = self.create_publication_for_autoload_report()

        sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=[
                {
                    "Id": linked_publication.row_id,
                    "AvitoId": "9999999999",
                    "status": "accepted",
                }
            ],
        )

        unlinked_publication = self.create_publication_for_autoload_report(
            row_id="SERVICE-ROW-002",
        )

        client = APIClient()
        client.force_authenticate(user=self.user)
        client.defaults["HTTP_X_WORKSPACE_ID"] = str(self.workspace.id)

        url = reverse(
            "avito-account-ads-list",
            kwargs={"avito_account_id": self.avito_account.id},
        )

        response = client.get(
            url,
            {
                "has_avito_id": "0",
                "page": 1,
                "page_size": 20,
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)

        results = response.json()["results"]

        self.assertTrue(results)
        self.assertTrue(
            all(item["has_avito_id"] is False for item in results)
        )
        self.assertEqual(
            [item["publication"] for item in results],
            [unlinked_publication.id],
        )

    def test_autoload_report_sync_api_links_publication_to_listing(self):
        publication = self.create_publication_for_autoload_report()

        client = APIClient()
        client.force_authenticate(user=self.user)
        client.defaults["HTTP_X_WORKSPACE_ID"] = str(self.workspace.id)

        url = reverse(
            "avito-account-autoload-report-sync",
            kwargs={"avito_account_id": self.avito_account.id},
        )

        response = client.post(
            url,
            {
                "report_rows": [
                    {
                        "Id": publication.row_id,
                        "AvitoId": "9999999999",
                        "status": "accepted",
                    }
                ]
            },
            format="json",
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)

        data = response.json()

        self.assertEqual(data["total_rows"], 1)
        self.assertEqual(data["accepted_rows"], 1)
        self.assertEqual(data["rejected_rows"], 0)
        self.assertEqual(data["linked_publications"], 1)
        self.assertEqual(data["created_listings"], 1)
        self.assertEqual(data["updated_listings"], 0)
        self.assertEqual(data["errors"], [])

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="9999999999",
        )

        self.assertEqual(listing.publication, publication)
        self.assertEqual(listing.row_id, publication.row_id)
        self.assertEqual(listing.source, AvitoListing.Source.SERVICE)
        self.assertEqual(listing.management_status, AvitoListing.ManagementStatus.MANAGED)

    def test_autoload_report_sync_updates_existing_publication_listing_with_new_avito_id(self):
        publication = self.create_publication_for_autoload_report()

        sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=[
                {
                    "Id": publication.row_id,
                    "AvitoId": "1111111111",
                    "status": "accepted",
                }
            ],
        )

        client = APIClient()
        client.force_authenticate(user=self.user)
        client.defaults["HTTP_X_WORKSPACE_ID"] = str(self.workspace.id)

        url = reverse(
            "avito-account-autoload-report-sync",
            kwargs={"avito_account_id": self.avito_account.id},
        )

        response = client.post(
            url,
            {
                "report_rows": [
                    {
                        "Id": publication.row_id,
                        "AvitoId": "2222222222",
                        "status": "accepted",
                    }
                ]
            },
            format="json",
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)

        data = response.json()

        self.assertEqual(data["created_listings"], 0)
        self.assertEqual(data["updated_listings"], 1)
        self.assertEqual(data["linked_publications"], 1)

        self.assertEqual(
            AvitoListing.objects.filter(
                workspace=self.workspace,
                avito_account=self.avito_account,
                publication=publication,
            ).count(),
            1,
        )

        listing = AvitoListing.objects.get(publication=publication)
        self.assertEqual(listing.avito_id, "2222222222")

    def create_publication_for_autoload_report(self, row_id="SERVICE-ROW-001"):
        batch = AdBatch.objects.create(
            workspace=self.workspace,
            source=AdBatch.Source.MANUAL,
            status=AdBatch.Status.COMPLETED,
            total_creatives=1,
            total_publications=1,
        )
        creative = AdCreative.objects.create(
            workspace=self.workspace,
            batch=batch,
            source=AdCreative.Source.MANUAL,
            title="Публикация из сервиса",
            description="<p>Описание из сервиса</p>",
            image_urls=["https://example.com/service-1.jpg"],
            base_data={
                "Category": "Ремонт и строительство",
                "Price": "1000",
            },
            option_data={
                "TargetAudience": "Частные лица и бизнес",
            },
        )

        return AdPublication.objects.create(
            workspace=self.workspace,
            avito_account=self.avito_account,
            creative=creative,
            batch=batch,
            source=AdPublication.Source.MANUAL,
            status=AdPublication.Status.ACTIVE,
            row_id=row_id,
            address="Москва",
        )

    def test_autoload_report_sync_links_publication_to_avito_listing(self):
        publication = self.create_publication_for_autoload_report()

        result = sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=[
                {
                    "Id": publication.row_id,
                    "AvitoId": "9999999999",
                    "status": "accepted",
                }
            ],
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.accepted_rows, 1)
        self.assertEqual(result.rejected_rows, 0)
        self.assertEqual(result.linked_publications, 1)
        self.assertEqual(result.created_listings, 1)
        self.assertEqual(result.updated_listings, 0)
        self.assertEqual(result.errors, [])

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="9999999999",
        )

        self.assertEqual(listing.publication, publication)
        self.assertEqual(listing.row_id, publication.row_id)
        self.assertEqual(listing.source, AvitoListing.Source.SERVICE)
        self.assertEqual(listing.management_status, AvitoListing.ManagementStatus.MANAGED)
        self.assertEqual(listing.title, "Публикация из сервиса")
        self.assertEqual(listing.description, "<p>Описание из сервиса</p>")
        self.assertEqual(listing.address, "Москва")
        self.assertEqual(listing.base_data["Price"], "1000")
        self.assertEqual(listing.option_data["TargetAudience"], "Частные лица и бизнес")

    def test_autoload_report_sync_is_idempotent_for_same_avito_id(self):
        publication = self.create_publication_for_autoload_report()

        report_rows = [
            {
                "Id": publication.row_id,
                "AvitoId": "9999999999",
                "status": "accepted",
            }
        ]

        first_result = sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=report_rows,
        )
        second_result = sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=report_rows,
        )

        self.assertEqual(first_result.created_listings, 1)
        self.assertEqual(second_result.created_listings, 0)
        self.assertEqual(second_result.updated_listings, 1)

        self.assertEqual(
            AvitoListing.objects.filter(
                workspace=self.workspace,
                avito_account=self.avito_account,
                avito_id="9999999999",
            ).count(),
            1,
        )

    def test_autoload_report_sync_marks_publication_error(self):
        publication = self.create_publication_for_autoload_report()

        result = sync_avito_autoload_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            report_rows=[
                {
                    "Id": publication.row_id,
                    "status": "rejected",
                    "error": "Не заполнено обязательное поле Price",
                }
            ],
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.accepted_rows, 0)
        self.assertEqual(result.rejected_rows, 1)
        self.assertEqual(result.created_listings, 0)

        publication.refresh_from_db()

        self.assertEqual(publication.status, AdPublication.Status.ERROR)
        self.assertEqual(
            publication.address_data["autoload_error"]["message"],
            "Не заполнено обязательное поле Price",
        )
        self.assertEqual(
            publication.address_data["autoload_error"]["status"],
            "rejected",
        )

    def test_csv_export_includes_managed_excel_listings(self):
        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(),
        )

        with tempfile.TemporaryDirectory() as output_dir:
            file_path = export_avito_account_publications_to_csv(
                workspace=self.workspace,
                avito_account=self.avito_account,
                output_dir=output_dir,
            )

            with open(file_path, newline="", encoding="utf-8") as csv_file:
                rows = list(csv.DictReader(csv_file, delimiter=";"))

        self.assertEqual(len(rows), 1)

        row = rows[0]
        self.assertEqual(row["Id"], "ROW-001")
        self.assertEqual(row["Title"], "Газосиликатные блоки")
        self.assertEqual(row["Address"], "Московская обл., Черноголовка")
        self.assertEqual(row["ContactPhone"], "74993919801")
        self.assertEqual(row["Price"], "4340")
        self.assertEqual(row["TargetAudience"], "Частные лица и бизнес")
        self.assertEqual(row["MinSaleQuantity"], "4")

    def test_csv_export_excludes_paused_excel_listings(self):
        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(),
        )

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="8036155996",
        )
        listing.desired_status = AvitoListing.DesiredStatus.PAUSE
        listing.save(update_fields=["desired_status", "updated_at"])

        with tempfile.TemporaryDirectory() as output_dir:
            file_path = export_avito_account_publications_to_csv(
                workspace=self.workspace,
                avito_account=self.avito_account,
                output_dir=output_dir,
            )

            with open(file_path, newline="", encoding="utf-8") as csv_file:
                rows = list(csv.DictReader(csv_file, delimiter=";"))

        self.assertEqual(rows, [])

    def setUp(self):
        self.user = User.objects.create_user(
            email="excel-import-owner@example.com",
            password="test",
        )
        self.workspace = Workspace.objects.create(
            name="Excel import workspace",
            slug="excel-import-workspace",
            owner=self.user,
        )
        WorkspaceMembership.objects.create(
            workspace=self.workspace,
            user=self.user,
            role=WorkspaceMembership.Role.OWNER,
            status=WorkspaceMembership.Status.ACTIVE,
        )
        self.avito_account = AvitoAccount.objects.create(
            workspace=self.workspace,
            name="Excel Avito account",
            external_account_id="excel-account-1",
        )

        ProductOptions.objects.create(
            option_title_ru="Целевая аудитория",
            option_title_en="TargetAudience",
        )
        ProductOptions.objects.create(
            option_title_ru="Минимальный заказ",
            option_title_en="MinSaleQuantity",
        )

    def build_excel_file(
            self,
            *,
            title="Газосиликатные блоки",
            price="4340",
            date_end="2026-06-01T10:00:00+03:00",
            unknown_required="Необязательный",
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Блоки-Газосиликат"

        headers = [
            "Уникальный идентификатор объявления",
            "Номер объявления на Авито",
            "Название объявления",
            "Описание объявления",
            "Ссылки на фото",
            "Адрес",
            "AvitoStatus",
            "AvitoDateEnd",
            "Номер телефона",
            "Цена",
            "Категория",
            "Целевая аудитория",
            "Минимальный заказ",
            "Неизвестная колонка",
        ]

        worksheet.cell(row=1, column=1).value = (
            "Для дома и дачи - Ремонт и строительство - Стройматериалы"
        )

        for index, header in enumerate(headers, start=1):
            worksheet.cell(row=2, column=index).value = header
            worksheet.cell(row=3, column=index).value = (
                unknown_required if header == "Неизвестная колонка" else "Необязательный"
            )
            worksheet.cell(row=4, column=index).value = "Подробнее о параметре"

        values = [
            "ROW-001",
            "8036155996",
            title,
            "<p>Описание объявления</p>",
            "https://example.com/1.jpg | https://example.com/2.jpg",
            "Московская обл., Черноголовка",
            "Активно",
            date_end,
            "74993919801",
            price,
            "Ремонт и строительство",
            "Частные лица и бизнес",
            "4",
            "не должно потеряться",
        ]

        for index, value in enumerate(values, start=1):
            worksheet.cell(row=5, column=index).value = value

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        stream.name = "avito_listings.xlsx"
        return stream

    def test_bulk_desired_status_updates_managed_excel_listings(self):
        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(),
        )

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="8036155996",
        )

        result = bulk_update_avito_listing_desired_status(
            workspace=self.workspace,
            avito_account=self.avito_account,
            listing_ids=[listing.id],
            desired_status=AvitoListing.DesiredStatus.PAUSE,
        )

        self.assertEqual(result["requested"], 1)
        self.assertEqual(result["matched"], 1)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(result["missing"], 0)
        self.assertEqual(result["desired_status"], "pause")

        listing.refresh_from_db()
        self.assertEqual(listing.desired_status, AvitoListing.DesiredStatus.PAUSE)

    def test_excel_import_preview_and_apply_api(self):
        client = APIClient()
        client.force_authenticate(user=self.user)
        client.defaults["HTTP_X_WORKSPACE_ID"] = str(self.workspace.id)

        preview_url = reverse(
            "avito-account-excel-import-preview",
            kwargs={"avito_account_id": self.avito_account.id},
        )

        preview_response = client.post(
            preview_url,
            {"file": self.build_excel_file()},
            format="multipart",
            HTTP_HOST="localhost",
        )

        self.assertEqual(preview_response.status_code, 200)
        preview_data = preview_response.json()

        self.assertEqual(preview_data["total_rows"], 1)
        self.assertEqual(preview_data["rows_with_errors"], 0)
        self.assertEqual(preview_data["rows"][0]["row_id"], "ROW-001")

        apply_url = reverse(
            "avito-account-excel-import-apply",
            kwargs={"avito_account_id": self.avito_account.id},
        )

        apply_response = client.post(
            apply_url,
            {"file": self.build_excel_file()},
            format="multipart",
            HTTP_HOST="localhost",
        )

        self.assertEqual(apply_response.status_code, 201)
        apply_data = apply_response.json()

        self.assertEqual(apply_data["total_rows"], 1)
        self.assertEqual(apply_data["skipped_rows"], 0)
        self.assertEqual(apply_data["created_listings"], 1)

        self.assertTrue(
            AvitoListing.objects.filter(
                workspace=self.workspace,
                avito_account=self.avito_account,
                avito_id="8036155996",
                source=AvitoListing.Source.AVITO_EXCEL,
            ).exists()
        )

    def test_avito_listing_api_filters_managed_excel_listings(self):
        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(),
        )

        client = APIClient()
        client.force_authenticate(user=self.user)
        client.defaults["HTTP_X_WORKSPACE_ID"] = str(self.workspace.id)

        url = reverse("avito-listing-api-list")

        response = client.get(
            url,
            {
                "source": "avito_excel",
                "management_status": "managed",
                "desired_status": "publish",
                "search": "Черноголовка",
                "page_size": 10,
            },
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()

        self.assertEqual(data["count"], 1)
        self.assertEqual(len(data["results"]), 1)

        listing = data["results"][0]
        self.assertEqual(listing["source"], "avito_excel")
        self.assertEqual(listing["management_status"], "managed")
        self.assertEqual(listing["desired_status"], "publish")
        self.assertEqual(listing["avito_id"], "8036155996")
        self.assertEqual(listing["row_id"], "ROW-001")

    def test_lifecycle_report_marks_listing_as_expires_soon(self):
        date_end = timezone.now() + timedelta(days=2)

        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(date_end=date_end.isoformat()),
        )

        report = build_avito_listing_lifecycle_report(
            workspace=self.workspace,
            avito_account=self.avito_account,
            soon_days=3,
        )

        self.assertEqual(report.total_checked, 1)
        self.assertEqual(report.expired, 0)
        self.assertEqual(report.expires_soon, 1)
        self.assertEqual(report.active_ok, 0)

        self.assertEqual(report.items[0].action, "expires_soon")
        self.assertEqual(report.items[0].avito_id, "8036155996")

    def test_bulk_management_status_updates_excel_listings(self):
        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(),
        )

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="8036155996",
        )

        result = bulk_update_avito_listing_management_status(
            workspace=self.workspace,
            avito_account=self.avito_account,
            listing_ids=[listing.id],
            management_status=AvitoListing.ManagementStatus.OUT_OF_SYNC,
        )

        self.assertEqual(result["requested"], 1)
        self.assertEqual(result["matched"], 1)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(result["missing"], 0)
        self.assertEqual(result["management_status"], "out_of_sync")

        listing.refresh_from_db()
        self.assertEqual(listing.management_status, AvitoListing.ManagementStatus.OUT_OF_SYNC)

    def test_preview_maps_excel_columns_to_internal_keys(self):
        result = preview_avito_excel_file(self.build_excel_file())

        self.assertEqual(result.total_sheets, 1)
        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.rows_with_errors, 0)

        row = result.rows[0]

        self.assertEqual(row.row_id, "ROW-001")
        self.assertEqual(row.avito_id, "8036155996")
        self.assertEqual(row.title, "Газосиликатные блоки")

        self.assertEqual(row.mapped_data["ContactPhone"], "74993919801")
        self.assertEqual(row.mapped_data["Price"], "4340")
        self.assertEqual(row.mapped_data["TargetAudience"], "Частные лица и бизнес")
        self.assertEqual(row.mapped_data["MinSaleQuantity"], "4")
        self.assertEqual(row.mapped_data["image_urls"], [
            "https://example.com/1.jpg",
            "https://example.com/2.jpg",
        ])

        self.assertEqual(row.unmapped_data, {})

    def test_preview_keeps_required_unknown_columns_as_unmapped(self):
        result = preview_avito_excel_file(
            self.build_excel_file(unknown_required="Обязательный")
        )

        row = result.rows[0]

        self.assertEqual(row.unmapped_data, {
            "Неизвестная колонка": "не должно потеряться",
        })

    def test_import_creates_managed_avito_listing_from_excel(self):
        result = import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(),
        )

        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.skipped_rows, 0)
        self.assertEqual(result.created_listings, 1)
        self.assertEqual(result.updated_listings, 0)

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="8036155996",
        )

        self.assertEqual(listing.source, AvitoListing.Source.AVITO_EXCEL)
        self.assertEqual(listing.management_status, AvitoListing.ManagementStatus.MANAGED)
        self.assertEqual(listing.desired_status, AvitoListing.DesiredStatus.PUBLISH)
        self.assertIsNone(listing.publication)

        self.assertEqual(listing.row_id, "ROW-001")
        self.assertEqual(listing.title, "Газосиликатные блоки")
        self.assertEqual(listing.address, "Московская обл., Черноголовка")
        self.assertEqual(listing.status, "Активно")
        self.assertEqual(listing.base_data["ContactPhone"], "74993919801")
        self.assertEqual(listing.base_data["Price"], "4340")
        self.assertEqual(listing.option_data["TargetAudience"], "Частные лица и бизнес")
        self.assertEqual(listing.option_data["MinSaleQuantity"], "4")
        self.assertEqual(listing.unmapped_data, {})

    def test_reimport_updates_existing_listing_without_duplicate(self):
        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(title="Старое название", price="4340"),
        )

        result = import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(title="Новое название", price="4500"),
        )

        self.assertEqual(result.created_listings, 0)
        self.assertEqual(result.updated_listings, 1)

        listings = AvitoListing.objects.filter(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="8036155996",
        )

        self.assertEqual(listings.count(), 1)

        listing = listings.get()
        self.assertEqual(listing.title, "Новое название")
        self.assertEqual(listing.base_data["Price"], "4500")

    def test_api_import_updates_only_observed_fields_for_excel_listing(self):
        import_avito_excel_file(
            workspace=self.workspace,
            avito_account=self.avito_account,
            file_obj=self.build_excel_file(),
        )

        listing = AvitoListing.objects.get(
            workspace=self.workspace,
            avito_account=self.avito_account,
            avito_id="8036155996",
        )

        old_title = listing.title
        old_description = listing.description
        old_base_data = dict(listing.base_data)
        old_option_data = dict(listing.option_data)

        updated_listing, was_created = upsert_avito_listing(
            self.avito_account,
            {
                "id": listing.avito_id,
                "status": "removed",
                "title": "API title must not overwrite Excel title",
                "url": "https://www.avito.ru/test",
            },
        )

        self.assertFalse(was_created)

        updated_listing.refresh_from_db()

        self.assertEqual(updated_listing.status, "removed")
        self.assertEqual(updated_listing.url, "https://www.avito.ru/test")
        self.assertEqual(updated_listing.title, old_title)
        self.assertEqual(updated_listing.description, old_description)
        self.assertEqual(updated_listing.base_data, old_base_data)
        self.assertEqual(updated_listing.option_data, old_option_data)
        self.assertEqual(
            updated_listing.imported_payload["api"]["title"],
            "API title must not overwrite Excel title",
        )


class AdScheduleCalculationTests(TestCase):
    timezone_name = "Europe/Moscow"
    tz = ZoneInfo(timezone_name)

    def dt(self, year, month, day, hour, minute):
        return datetime(year, month, day, hour, minute, tzinfo=self.tz)

    def test_next_run_is_today_when_slot_is_still_in_future(self):
        result = calculate_next_run_at(
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            anchor_date=date(2026, 5, 4),
            now=self.dt(2026, 5, 4, 9, 59),
            timezone_name=self.timezone_name,
        )

        self.assertEqual(result, self.dt(2026, 5, 4, 10, 0))

    def test_next_run_moves_to_next_valid_week_when_slot_already_passed(self):
        result = calculate_next_run_at(
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            anchor_date=date(2026, 5, 4),
            now=self.dt(2026, 5, 4, 10, 1),
            timezone_name=self.timezone_name,
        )

        self.assertEqual(result, self.dt(2026, 5, 11, 10, 0))

    def test_thursday_schedule_does_not_run_on_wednesday(self):
        result = calculate_next_run_at(
            schedule={
                "frequency": 1,
                "days": [None, None, None, "14:30", None, None, None],
            },
            anchor_date=date(2026, 5, 4),
            now=self.dt(2026, 5, 6, 14, 30),
            timezone_name=self.timezone_name,
        )

        self.assertEqual(result, self.dt(2026, 5, 7, 14, 30))

    def test_frequency_two_skips_non_matching_iso_week(self):
        result = calculate_next_run_at(
            schedule={
                "frequency": 2,
                "days": ["10:00", None, None, None, None, None, None],
            },
            anchor_date=date(2026, 5, 4),
            now=self.dt(2026, 5, 11, 10, 0),
            timezone_name=self.timezone_name,
        )

        self.assertEqual(result, self.dt(2026, 5, 18, 10, 0))

    def test_frequency_four_returns_slot_after_28_days(self):
        result = calculate_next_run_at(
            schedule={
                "frequency": 4,
                "days": ["10:00", None, None, None, None, None, None],
            },
            anchor_date=date(2026, 5, 4),
            now=self.dt(2026, 5, 4, 10, 1),
            timezone_name=self.timezone_name,
        )

        self.assertEqual(result, self.dt(2026, 6, 1, 10, 0))

    def test_schedule_requires_valid_frequency(self):
        with self.assertRaisesMessage(AdScheduleError, "frequency должен быть 1, 2, 3 или 4"):
            normalize_schedule({
                "frequency": 5,
                "days": ["10:00", None, None, None, None, None, None],
            })

    def test_schedule_requires_seven_days(self):
        with self.assertRaisesMessage(AdScheduleError, "days должен быть массивом из 7 элементов"):
            normalize_schedule({
                "frequency": 1,
                "days": ["10:00"],
            })

    def test_schedule_requires_strict_hh_mm_time(self):
        with self.assertRaisesMessage(AdScheduleError, "Время в расписании должно быть строго в формате HH:mm"):
            normalize_schedule({
                "frequency": 1,
                "days": ["9:00", None, None, None, None, None, None],
            })

    def test_schedule_requires_at_least_one_selected_day(self):
        with self.assertRaisesMessage(AdScheduleError, "В расписании должен быть выбран хотя бы один день"):
            normalize_schedule({
                "frequency": 1,
                "days": [None, None, None, None, None, None, None],
            })


class AdGenerationServiceTests(TestCase):
    def create_image_asset(self, *, workspace, url, uploaded_by=None):
        original_filename = url.rstrip("/").rsplit("/", 1)[-1] or "image.jpg"

        return AdImageAsset.objects.create(
            workspace=workspace,
            uploaded_by=uploaded_by or workspace.owner,
            image=f"tests/{original_filename}",
            url=url,
            original_filename=original_filename,
            content_type="image/jpeg",
            size_bytes=1,
            checksum=url,
        )

    def attach_task_images(self, task, *, main_urls, additional_urls=None):
        main_assets = [
            self.create_image_asset(workspace=task.workspace, url=url)
            for url in main_urls
        ]
        additional_assets = [
            self.create_image_asset(workspace=task.workspace, url=url)
            for url in additional_urls or []
        ]

        task.main_image_assets.set(main_assets)
        task.additional_image_assets.set(additional_assets)

        return main_assets, additional_assets

    def test_legacy_update_product_price_task_is_disabled(self):
        from avitotask.tasks import update_product_price

        result = update_product_price(123456)

        self.assertEqual(result["status"], "disabled")
        self.assertEqual(result["replacement"], "AdGenerationTask")

    def test_products_api_does_not_create_legacy_product_after_full_flow(self):
        user = User.objects.create_user(email="no-legacy-product-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="No legacy product workspace",
            slug="no-legacy-product-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="No Legacy Account")
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)
        main_asset = self.create_image_asset(
            workspace=workspace,
            url="https://example.com/no-legacy.jpg",
            uploaded_by=user,
        )

        create_response = client.post(
            reverse("product-api-list"),
            {
                "name": "No legacy task",
                "activate": False,
                "price": 1000,
                "titles": ["No legacy title"],
                "descriptions": ["No legacy description {{ TITLE }} / {{ SKU }}"],
                "main_image_asset_ids": [main_asset.id],
                "addresses": ["No Legacy Address"],
                "schedule": {
                    "frequency": 1,
                    "days": ["10:00", None, None, None, None, None, None],
                },
                "schedule_anchor_date": "2026-05-04",
                "schedule_timezone": "Europe/Moscow",
                "avito_account_ids": [account.id],
            },
            format="json",
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        self.assertEqual(create_response.status_code, 201)

        task_id = create_response.data["id"]

        random_response = client.post(
            reverse("product-random-api", args=[task_id]),
            {},
            format="json",
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        self.assertEqual(random_response.status_code, 200)

        self.assertEqual(AdGenerationTask.objects.filter(workspace=workspace).count(), 1)
        self.assertEqual(AdCreative.objects.filter(workspace=workspace, task_id=task_id).count(), 1)
        self.assertEqual(AdPublication.objects.filter(workspace=workspace, task_id=task_id).count(), 1)

    def test_products_api_patch_recalculates_next_update_time_when_schedule_changes(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="products-api-patch-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Products API patch workspace",
            slug="products-api-patch-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Patch schedule task",
            is_active=True,
            titles=["Patch title"],
            descriptions={"0": "Patch description"},
            addresses=["Patch Address"],
            price=100,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=datetime(2026, 5, 4, 10, 0, tzinfo=tz),
        )
        self.attach_task_images(task, main_urls=["https://example.com/patch.jpg"])

        with patch(
                "avitotask.services.ad_schedule.timezone.now",
                return_value=datetime(2026, 5, 6, 12, 0, tzinfo=tz),
        ):
            response = client.patch(
                reverse("product-api-detail", args=[task.id]),
                {
                    "schedule": {
                        "frequency": 2,
                        "days": [None, None, None, "14:30", None, None, None],
                    },
                    "schedule_anchor_date": "2026-05-04",
                    "next_update_time": "2030-01-01T00:00:00+03:00",
                    "last_run_status": "success",
                    "last_run_error": "frontend must not set this",
                },
                format="json",
                HTTP_X_WORKSPACE_ID=str(workspace.id),
            )

        task.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(task.schedule, {
            "frequency": 2,
            "days": [None, None, None, "14:30", None, None, None],
        })
        self.assertEqual(task.publication_interval_days, 14)
        self.assertEqual(
            task.next_update_time.astimezone(tz),
            datetime(2026, 5, 7, 14, 30, tzinfo=tz),
        )
        self.assertEqual(task.last_run_status, AdGenerationTask.LastRunStatus.IDLE)
        self.assertIsNone(task.last_run_error)

        self.assertEqual(response.data["next_update_time"], "2026-05-07T14:30:00+03:00")
        self.assertEqual(response.data["last_run_status"], "idle")
        self.assertIsNone(response.data["last_run_error"])

    def test_export_avito_account_csv_task_skips_account_that_is_already_exporting(self):
        user = User.objects.create_user(email="csv-export-lock-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="CSV export lock workspace",
            slug="csv-export-lock-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="CSV Export Lock Account",
            export_status=AvitoAccount.ExportStatus.EXPORTING,
        )

        with patch("avitotask.tasks.export_avito_account_publications_to_csv") as export_mock:
            result = export_avito_account_csv_task(account.id)

        account.refresh_from_db()

        self.assertEqual(result["status"], "skipped")
        self.assertEqual(result["reason"], "already_exporting")
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.EXPORTING)
        export_mock.assert_not_called()

    def test_products_api_patch_recalculates_next_update_time_when_anchor_changes(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="products-api-anchor-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Products API anchor workspace",
            slug="products-api-anchor-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Patch anchor task",
            is_active=True,
            schedule={
                "frequency": 2,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=datetime(2026, 5, 18, 10, 0, tzinfo=tz),
        )

        with patch(
                "avitotask.services.ad_schedule.timezone.now",
                return_value=datetime(2026, 5, 11, 9, 0, tzinfo=tz),
        ):
            response = client.patch(
                reverse("product-api-detail", args=[task.id]),
                {
                    "schedule_anchor_date": "2026-05-11",
                },
                format="json",
                HTTP_X_WORKSPACE_ID=str(workspace.id),
            )

        task.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(task.schedule_anchor_date, date(2026, 5, 11))
        self.assertEqual(task.next_update_time, datetime(2026, 5, 11, 10, 0, tzinfo=tz))
        self.assertEqual(response.data["next_update_time"], "2026-05-11T10:00:00+03:00")

    def test_products_api_patch_rejects_invalid_schedule_without_changing_task(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="products-api-invalid-patch-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Products API invalid patch workspace",
            slug="products-api-invalid-patch-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Invalid patch task",
            is_active=True,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=datetime(2026, 5, 4, 10, 0, tzinfo=tz),
        )

        response = client.patch(
            reverse("product-api-detail", args=[task.id]),
            {
                "schedule": {
                    "frequency": 5,
                    "days": ["10:00", None, None, None, None, None, None],
                },
            },
            format="json",
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        task.refresh_from_db()

        self.assertEqual(response.status_code, 400)
        self.assertIn("schedule", response.data)
        self.assertEqual(task.schedule, {
            "frequency": 1,
            "days": ["10:00", None, None, None, None, None, None],
        })
        self.assertEqual(task.next_update_time, datetime(2026, 5, 4, 10, 0, tzinfo=tz))

    def test_run_autogeneration_task_queues_csv_export_tasks_after_commit(self):
        user = User.objects.create_user(email="runner-export-queue-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Runner export queue workspace",
            slug="runner-export-queue-workspace",
            owner=user,
        )
        account_1 = AvitoAccount.objects.create(workspace=workspace, name="Queue Account 1")
        account_2 = AvitoAccount.objects.create(workspace=workspace, name="Queue Account 2")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Export queue task",
            is_active=False,
            titles=["Export queue title"],
            descriptions={"0": "Export queue description {{ TITLE }} / {{ SKU }}"},
            addresses=["Export Queue Address"],
            price=100,
        )
        self.attach_task_images(task, main_urls=["https://example.com/export-queue.jpg"])
        task.avito_accounts.add(account_1, account_2)

        with patch("avitotask.tasks.export_avito_account_csv_task.delay") as delay_mock:
            with self.captureOnCommitCallbacks(execute=True):
                result = run_autogeneration_task(
                    task.id,
                    triggered_by="manual",
                    user=user,
                )

        account_1.refresh_from_db()
        account_2.refresh_from_db()

        self.assertEqual(result.csv_export_status, "queued")
        self.assertEqual(account_1.export_status, AvitoAccount.ExportStatus.QUEUED)
        self.assertEqual(account_2.export_status, AvitoAccount.ExportStatus.QUEUED)
        delay_mock.assert_has_calls([
            call(account_1.id),
            call(account_2.id),
        ], any_order=True)
        self.assertEqual(delay_mock.call_count, 2)

    def test_duplicate_scheduled_run_does_not_queue_csv_export_again(self):
        tz = ZoneInfo("Europe/Moscow")
        scheduled_for = datetime(2026, 5, 4, 10, 0, tzinfo=tz)

        user = User.objects.create_user(email="runner-export-duplicate-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Runner export duplicate workspace",
            slug="runner-export-duplicate-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Queue Duplicate Account")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Export duplicate task",
            is_active=True,
            titles=["Export duplicate title"],
            descriptions={"0": "Export duplicate description {{ TITLE }} / {{ SKU }}"},
            addresses=["Export Duplicate Address"],
            price=100,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=scheduled_for,
        )
        self.attach_task_images(task, main_urls=["https://example.com/export-duplicate.jpg"])
        task.avito_accounts.add(account)

        with patch("avitotask.tasks.export_avito_account_csv_task.delay") as delay_mock:
            with self.captureOnCommitCallbacks(execute=True):
                first_result = run_autogeneration_task(
                    task.id,
                    triggered_by="schedule",
                    scheduled_for=scheduled_for,
                    now=scheduled_for,
                )

            delay_mock.reset_mock()

            with self.captureOnCommitCallbacks(execute=True):
                second_result = run_autogeneration_task(
                    task.id,
                    triggered_by="schedule",
                    scheduled_for=scheduled_for,
                    now=scheduled_for,
                )

        self.assertTrue(first_result.created)
        self.assertFalse(second_result.created)
        delay_mock.assert_not_called()
        self.assertEqual(AdGenerationTaskRun.objects.filter(task=task).count(), 1)
        self.assertEqual(AdCreative.objects.filter(task=task).count(), 1)

    def test_scheduler_uses_autogeneration_pipeline_and_creates_task_run(self):
        tz = ZoneInfo("Europe/Moscow")
        scheduled_for = datetime(2026, 5, 4, 10, 0, tzinfo=tz)

        user = User.objects.create_user(email="scheduler-runner-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Scheduler runner workspace",
            slug="scheduler-runner-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Scheduler Runner Account")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Scheduler runner task",
            is_active=True,
            titles=["Scheduler runner title"],
            descriptions={"0": "Scheduler runner description {{ TITLE }} / {{ SKU }}"},
            addresses=["Scheduler Runner Address"],
            price=100,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=scheduled_for,
        )
        self.attach_task_images(task, main_urls=["https://example.com/scheduler-runner.jpg"])
        task.avito_accounts.add(account)

        generated_count = run_due_ad_generation_tasks(
            now_dt=scheduled_for,
        )

        task.refresh_from_db()

        self.assertEqual(generated_count, 1)
        self.assertEqual(AdGenerationTaskRun.objects.filter(task=task).count(), 1)

        run = AdGenerationTaskRun.objects.get(task=task)
        self.assertEqual(run.triggered_by, AdGenerationTaskRun.TriggeredBy.SCHEDULE)
        self.assertEqual(run.scheduled_for, scheduled_for)
        self.assertEqual(run.status, AdGenerationTaskRun.Status.SUCCESS)
        self.assertEqual(run.publications_count, 1)

        self.assertEqual(AdCreative.objects.filter(task=task).count(), 1)
        self.assertEqual(AdPublication.objects.filter(task=task).count(), 1)
        self.assertEqual(task.last_run_status, AdGenerationTask.LastRunStatus.SUCCESS)
        self.assertEqual(task.last_successful_run_at, scheduled_for)
        self.assertEqual(task.next_update_time, datetime(2026, 5, 11, 10, 0, tzinfo=tz))

    def test_scheduler_does_not_run_frequency_two_on_non_matching_week(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="scheduler-frequency-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Scheduler frequency workspace",
            slug="scheduler-frequency-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Scheduler Frequency Account")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Scheduler frequency task",
            is_active=True,
            titles=["Frequency title"],
            descriptions={"0": "Frequency description {{ TITLE }} / {{ SKU }}"},
            addresses=["Frequency Address"],
            price=100,
            schedule={
                "frequency": 2,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            # Специально ставим неверную due дату, чтобы scheduler не запускал
            # только потому, что next_update_time <= now.
            next_update_time=datetime(2026, 5, 11, 10, 0, tzinfo=tz),
        )
        self.attach_task_images(task, main_urls=["https://example.com/frequency.jpg"])
        task.avito_accounts.add(account)

        generated_count = run_due_ad_generation_tasks(
            now_dt=datetime(2026, 5, 11, 10, 0, tzinfo=tz),
        )

        task.refresh_from_db()

        self.assertEqual(generated_count, 0)
        self.assertEqual(AdGenerationTaskRun.objects.filter(task=task).count(), 0)
        self.assertEqual(AdCreative.objects.filter(task=task).count(), 0)
        self.assertEqual(task.next_update_time, datetime(2026, 5, 18, 10, 0, tzinfo=tz))

    def test_scheduler_duplicate_tick_does_not_create_second_creative(self):
        tz = ZoneInfo("Europe/Moscow")
        scheduled_for = datetime(2026, 5, 4, 10, 0, tzinfo=tz)

        user = User.objects.create_user(email="scheduler-duplicate-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Scheduler duplicate workspace",
            slug="scheduler-duplicate-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Scheduler Duplicate Account")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Scheduler duplicate task",
            is_active=True,
            titles=["Scheduler duplicate title"],
            descriptions={"0": "Scheduler duplicate description {{ TITLE }} / {{ SKU }}"},
            addresses=["Scheduler Duplicate Address"],
            price=100,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=scheduled_for,
        )
        self.attach_task_images(task, main_urls=["https://example.com/scheduler-duplicate.jpg"])
        task.avito_accounts.add(account)

        first_count = run_due_ad_generation_tasks(now_dt=scheduled_for)

        # Имитируем повторный tick с тем же scheduled_for.
        AdGenerationTask.objects.filter(id=task.id).update(next_update_time=scheduled_for)

        second_count = run_due_ad_generation_tasks(now_dt=scheduled_for)

        self.assertEqual(first_count, 1)
        self.assertEqual(second_count, 0)
        self.assertEqual(AdGenerationTaskRun.objects.filter(task=task).count(), 1)
        self.assertEqual(AdCreative.objects.filter(task=task).count(), 1)
        self.assertEqual(AdPublication.objects.filter(task=task).count(), 1)

    def test_run_autogeneration_task_manual_creates_run_creative_publications_and_dirty_export(self):
        user = User.objects.create_user(email="runner-manual-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Runner manual workspace",
            slug="runner-manual-workspace",
            owner=user,
        )
        account_1 = AvitoAccount.objects.create(workspace=workspace, name="Runner Account 1")
        account_2 = AvitoAccount.objects.create(workspace=workspace, name="Runner Account 2")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Manual runner task",
            is_active=False,
            titles=["Runner title"],
            descriptions={"0": "Runner description {{ TITLE }} / {{ SKU }}"},
            addresses=["Runner Address 1", "Runner Address 2"],
            price=100,
        )
        self.attach_task_images(task, main_urls=["https://example.com/runner.jpg"])
        task.avito_accounts.add(account_1, account_2)

        result = run_autogeneration_task(
            task.id,
            triggered_by="manual",
            user=user,
        )

        task.refresh_from_db()
        account_1.refresh_from_db()
        account_2.refresh_from_db()

        self.assertEqual(result.run.status, AdGenerationTaskRun.Status.SUCCESS)
        self.assertEqual(result.run.triggered_by, AdGenerationTaskRun.TriggeredBy.MANUAL)
        self.assertEqual(result.run.creative_id, result.creative.id)
        self.assertEqual(result.run.batch_id, result.batch.id)
        self.assertEqual(result.run.publications_count, 4)
        self.assertEqual(result.publications_count, 4)

        self.assertEqual(AdCreative.objects.filter(task=task).count(), 1)
        self.assertEqual(AdPublication.objects.filter(task=task).count(), 4)
        self.assertEqual(account_1.export_status, AvitoAccount.ExportStatus.DIRTY)
        self.assertEqual(account_2.export_status, AvitoAccount.ExportStatus.DIRTY)

        self.assertEqual(task.last_run_status, AdGenerationTask.LastRunStatus.SUCCESS)
        self.assertIsNotNone(task.last_run_at)
        self.assertEqual(task.last_successful_run_at, task.last_run_at)
        self.assertIsNone(task.last_run_error)

    def test_run_autogeneration_task_schedule_prevents_duplicate_scheduled_run(self):
        tz = ZoneInfo("Europe/Moscow")
        scheduled_for = datetime(2026, 5, 4, 10, 0, tzinfo=tz)

        user = User.objects.create_user(email="runner-duplicate-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Runner duplicate workspace",
            slug="runner-duplicate-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Runner Duplicate Account")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Duplicate runner task",
            is_active=True,
            titles=["Duplicate title"],
            descriptions={"0": "Duplicate description {{ TITLE }} / {{ SKU }}"},
            addresses=["Duplicate Address"],
            price=100,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=scheduled_for,
        )
        self.attach_task_images(task, main_urls=["https://example.com/duplicate.jpg"])
        task.avito_accounts.add(account)

        first_result = run_autogeneration_task(
            task.id,
            triggered_by="schedule",
            scheduled_for=scheduled_for,
            now=scheduled_for,
        )
        second_result = run_autogeneration_task(
            task.id,
            triggered_by="schedule",
            scheduled_for=scheduled_for,
            now=scheduled_for,
        )

        self.assertEqual(first_result.run.id, second_result.run.id)
        self.assertEqual(AdGenerationTaskRun.objects.filter(task=task).count(), 1)
        self.assertEqual(AdCreative.objects.filter(task=task).count(), 1)
        self.assertEqual(AdPublication.objects.filter(task=task).count(), 1)

    def test_run_autogeneration_task_schedule_records_error_for_inactive_task(self):
        tz = ZoneInfo("Europe/Moscow")
        scheduled_for = datetime(2026, 5, 4, 10, 0, tzinfo=tz)

        user = User.objects.create_user(email="runner-inactive-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Runner inactive workspace",
            slug="runner-inactive-workspace",
            owner=user,
        )

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Inactive schedule runner task",
            is_active=False,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
        )

        with self.assertRaisesMessage(AdGenerationError, "Задача генерации не активна"):
            run_autogeneration_task(
                task.id,
                triggered_by="schedule",
                scheduled_for=scheduled_for,
                now=scheduled_for,
            )

        task.refresh_from_db()
        run = AdGenerationTaskRun.objects.get(task=task)

        self.assertEqual(run.status, AdGenerationTaskRun.Status.ERROR)
        self.assertEqual(run.error, "Задача генерации не активна")
        self.assertEqual(task.last_run_status, AdGenerationTask.LastRunStatus.ERROR)
        self.assertEqual(task.last_run_error, "Задача генерации не активна")

    def create_api_client_for_workspace(self, *, user, workspace):
        WorkspaceMembership.objects.create(
            workspace=workspace,
            user=user,
            role=WorkspaceMembership.Role.OWNER,
            status=WorkspaceMembership.Status.ACTIVE,
        )

        client = APIClient()
        client.force_authenticate(user=user)

        return client

    def test_generate_ads_from_task_creates_publications_for_each_address_and_account(self):
        user = User.objects.create_user(email="owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Test workspace",
            slug="test-workspace",
            owner=user,
        )

        account_1 = AvitoAccount.objects.create(workspace=workspace, name="Account 1")
        account_2 = AvitoAccount.objects.create(workspace=workspace, name="Account 2")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Test task",
            is_active=True,
            titles=["Title 1"],
            descriptions={"1": "Description {{ TITLE }} / {{ SKU }}"},
            addresses=["Address 1", "Address 2"],
            price=100,
        )
        self.attach_task_images(task, main_urls=["https://example.com/main.jpg"])
        task.avito_accounts.add(account_1, account_2)

        result = generate_ads_from_task(task.id, workspace=workspace)

        publications = AdPublication.objects.filter(
            workspace=workspace,
            creative=result.creative,
        )

        self.assertEqual(publications.count(), 4)
        self.assertEqual(result.batch.total_creatives, 1)
        self.assertEqual(result.batch.total_publications, 4)
        self.assertEqual(len(result.publications), 4)
        self.assertEqual(
            set(publications.values_list("address", flat=True)),
            {"Address 1", "Address 2"},
        )
        self.assertEqual(
            set(publications.values_list("avito_account__name", flat=True)),
            {"Account 1", "Account 2"},
        )

    def test_create_manual_mass_posting_creates_creative_and_publications(self):
        user = User.objects.create_user(email="manual-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Manual workspace",
            slug="manual-workspace",
            owner=user,
        )

        account_1 = AvitoAccount.objects.create(workspace=workspace, name="Manual Account 1")
        account_2 = AvitoAccount.objects.create(workspace=workspace, name="Manual Account 2")

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account_1, account_2],
            addresses=["Manual Address 1", "Manual Address 2", "Manual Address 3"],
            title="Ручное объявление",
            description="Описание ручного объявления",
            image_urls=["https://example.com/manual-main.jpg"],
            base_data={"Price": 500, "Category": "Ремонт и строительство"},
            option_data={"Condition": "Новое"},
        )

        publications = AdPublication.objects.filter(
            workspace=workspace,
            creative=result.creative,
        )

        self.assertEqual(result.batch.source, "manual")
        self.assertEqual(result.creative.source, "manual")
        self.assertIsNone(result.creative.task)
        self.assertEqual(publications.count(), 6)
        self.assertEqual(len(result.publications), 6)
        self.assertEqual(result.batch.total_creatives, 1)
        self.assertEqual(result.batch.total_publications, 6)
        self.assertEqual(
            set(publications.values_list("address", flat=True)),
            {"Manual Address 1", "Manual Address 2", "Manual Address 3"},
        )

    def test_update_ad_publication_changes_only_selected_publication_overrides(self):
        user = User.objects.create_user(email="edit-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Edit workspace",
            slug="edit-workspace",
            owner=user,
        )

        account = AvitoAccount.objects.create(workspace=workspace, name="Edit Account")

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Edit Address 1", "Edit Address 2"],
            title="Editable ad",
            description="Editable description",
            image_urls=["https://example.com/edit.jpg"],
            base_data={"Price": 500, "Category": "Стройматериалы"},
            option_data={},
        )

        first_publication = result.publications[0]
        second_publication = result.publications[1]

        updated_publication = update_ad_publication(
            publication_id=first_publication.id,
            workspace=workspace,
            overrides={"Price": 700, "Title": "Индивидуальный заголовок"},
            address="Updated Address 1",
        )

        second_publication.refresh_from_db()

        self.assertEqual(updated_publication.overrides["Price"], 700)
        self.assertEqual(updated_publication.overrides["Title"], "Индивидуальный заголовок")
        self.assertEqual(updated_publication.address, "Updated Address 1")

        self.assertEqual(second_publication.overrides, {})
        self.assertEqual(second_publication.address, "Edit Address 2")
        self.assertEqual(result.creative.base_data["Price"], 500)
        self.assertEqual(result.creative.title, "Editable ad")

    def test_update_ad_creative_updates_shared_data_and_clears_matching_publication_overrides(self):
        user = User.objects.create_user(email="bulk-edit-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Bulk edit workspace",
            slug="bulk-edit-workspace",
            owner=user,
        )

        account = AvitoAccount.objects.create(workspace=workspace, name="Bulk Edit Account")

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Bulk Address 1", "Bulk Address 2"],
            title="Old shared title",
            description="Old shared description",
            image_urls=["https://example.com/old.jpg"],
            base_data={"Price": 500, "Category": "Стройматериалы"},
            option_data={"Condition": "Новое"},
        )

        first_publication = result.publications[0]
        second_publication = result.publications[1]

        update_ad_publication(
            publication_id=first_publication.id,
            workspace=workspace,
            overrides={"Price": 700, "Title": "Individual title"},
        )

        updated_creative = update_ad_creative(
            creative_id=result.creative.id,
            workspace=workspace,
            title="New shared title",
            base_data={"Price": 900},
            option_data={"Condition": "Б/у"},
            clear_publication_override_fields=["Price", "Title", "Condition"],
        )

        first_publication.refresh_from_db()
        second_publication.refresh_from_db()

        self.assertEqual(updated_creative.title, "New shared title")
        self.assertEqual(updated_creative.base_data["Price"], 900)
        self.assertEqual(updated_creative.base_data["Category"], "Стройматериалы")
        self.assertEqual(updated_creative.option_data["Condition"], "Б/у")

        self.assertNotIn("Price", first_publication.overrides)
        self.assertNotIn("Title", first_publication.overrides)
        self.assertEqual(second_publication.overrides, {})

    def test_build_publication_export_row_merges_creative_data_and_publication_overrides(self):
        user = User.objects.create_user(email="export-row-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Export row workspace",
            slug="export-row-workspace",
            owner=user,
        )

        account = AvitoAccount.objects.create(workspace=workspace, name="Export Row Account")

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Original Address"],
            title="Shared title",
            description="Shared description",
            image_urls=["https://example.com/1.jpg", "https://example.com/2.jpg"],
            base_data={"Price": 500, "Category": "Стройматериалы"},
            option_data={"Condition": "Новое"},
        )

        publication = update_ad_publication(
            publication_id=result.publications[0].id,
            workspace=workspace,
            overrides={
                "Price": 700,
                "Title": "Individual title",
                "CustomField": "Custom value",
            },
            address="Updated Address",
        )

        row = build_publication_export_row(publication)

        self.assertEqual(row["Id"], publication.row_id)
        self.assertEqual(row["Title"], "Individual title")
        self.assertEqual(row["Description"], "Shared description")
        self.assertEqual(row["ImageUrls"], "https://example.com/1.jpg | https://example.com/2.jpg")
        self.assertEqual(row["Address"], "Updated Address")
        self.assertEqual(row["Price"], 700)
        self.assertEqual(row["Category"], "Стройматериалы")
        self.assertEqual(row["Condition"], "Новое")
        self.assertEqual(row["CustomField"], "Custom value")

    def test_export_avito_account_publications_to_csv_writes_active_publications(self):
        user = User.objects.create_user(email="csv-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="CSV workspace",
            slug="csv-workspace",
            owner=user,
        )

        account = AvitoAccount.objects.create(workspace=workspace, name="CSV Account")

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["CSV Address 1", "CSV Address 2"],
            title="CSV title",
            description="CSV description",
            image_urls=["https://example.com/csv.jpg"],
            base_data={"Price": 500, "Category": "Стройматериалы"},
            option_data={"Condition": "Новое"},
        )

        update_ad_publication(
            publication_id=result.publications[0].id,
            workspace=workspace,
            overrides={"Price": 700},
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = export_avito_account_publications_to_csv(
                workspace=workspace,
                avito_account=account,
                output_dir=Path(temp_dir),
            )

            with open(file_path, newline="", encoding="utf-8") as csv_file:
                rows = list(csv.DictReader(csv_file, delimiter=";"))

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["Title"], "CSV title")
        self.assertEqual(rows[0]["Address"], "CSV Address 1")
        self.assertEqual(rows[0]["Price"], "700")
        self.assertEqual(rows[1]["Price"], "500")
        self.assertIn("Condition", rows[0])

    def test_csv_export_uses_backend_approved_columns_only(self):
        user = User.objects.create_user(email="csv-columns-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="CSV columns workspace",
            slug="csv-columns-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="CSV Columns Account")
        ProductOptions.objects.create(
            option_title_ru="Backend approved",
            option_title_en="BackendApproved",
        )

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["CSV Columns Address"],
            title="CSV columns title",
            description="CSV columns description",
            image_urls=["https://example.com/csv-columns.jpg"],
            base_data={
                "Price": 500,
                "Category": "Стройматериалы",
                "FrontendBaseOnly": "must not leak",
            },
            option_data={
                "BackendApproved": "yes",
                "FrontendOnly": "must not leak",
            },
        )

        update_ad_publication(
            publication_id=result.publications[0].id,
            workspace=workspace,
            overrides={
                "Title": "CSV override title",
                "FrontendOverrideOnly": "must not leak",
            },
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = export_avito_account_publications_to_csv(
                workspace=workspace,
                avito_account=account,
                output_dir=Path(temp_dir),
            )

            with open(file_path, newline="", encoding="utf-8") as csv_file:
                reader = csv.DictReader(csv_file, delimiter=";")
                rows = list(reader)
                fieldnames = reader.fieldnames

        self.assertEqual(len(rows), 1)

        self.assertIn("Id", fieldnames)
        self.assertIn("Title", fieldnames)
        self.assertIn("Description", fieldnames)
        self.assertIn("ImageUrls", fieldnames)
        self.assertIn("Address", fieldnames)
        self.assertIn("Price", fieldnames)
        self.assertIn("Category", fieldnames)
        self.assertIn("BackendApproved", fieldnames)

        self.assertNotIn("FrontendBaseOnly", fieldnames)
        self.assertNotIn("FrontendOnly", fieldnames)
        self.assertNotIn("FrontendOverrideOnly", fieldnames)

        self.assertEqual(rows[0]["Title"], "CSV override title")
        self.assertEqual(rows[0]["BackendApproved"], "yes")
        self.assertNotIn("FrontendBaseOnly", rows[0])
        self.assertNotIn("FrontendOnly", rows[0])
        self.assertNotIn("FrontendOverrideOnly", rows[0])

    def test_csv_export_writes_backend_headers_even_without_publications(self):
        user = User.objects.create_user(email="csv-empty-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="CSV empty workspace",
            slug="csv-empty-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="CSV Empty Account",
            export_status=AvitoAccount.ExportStatus.DIRTY,
        )

        ProductOptions.objects.create(
            option_title_ru="Backend approved empty",
            option_title_en="BackendApprovedEmpty",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = export_avito_account_publications_to_csv(
                workspace=workspace,
                avito_account=account,
                output_dir=Path(temp_dir),
            )

            with open(file_path, newline="", encoding="utf-8") as csv_file:
                reader = csv.DictReader(csv_file, delimiter=";")
                rows = list(reader)
                fieldnames = reader.fieldnames

        account.refresh_from_db()

        self.assertEqual(rows, [])
        self.assertIn("Id", fieldnames)
        self.assertIn("Title", fieldnames)
        self.assertIn("BackendApprovedEmpty", fieldnames)
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.CLEAN)
        self.assertTrue(account.export_file_path)

    def test_account_export_status_changes_after_publication_change_and_export(self):
        user = User.objects.create_user(email="dirty-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Dirty workspace",
            slug="dirty-workspace",
            owner=user,
        )

        account = AvitoAccount.objects.create(workspace=workspace, name="Dirty Account")

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Dirty Address"],
            title="Dirty title",
            description="Dirty description",
            image_urls=["https://example.com/dirty.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        account.refresh_from_db()
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.DIRTY)

        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = export_avito_account_publications_to_csv(
                workspace=workspace,
                avito_account=account,
                output_dir=Path(temp_dir),
            )

        account.refresh_from_db()
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.CLEAN)
        self.assertEqual(account.export_file_path, str(file_path))
        self.assertIsNotNone(account.last_exported_at)

        update_ad_publication(
            publication_id=result.publications[0].id,
            workspace=workspace,
            overrides={"Price": 700},
        )

        account.refresh_from_db()
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.DIRTY)

    def test_export_dirty_avito_accounts_csv_task_exports_dirty_accounts(self):
        user = User.objects.create_user(email="celery-export-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Celery export workspace",
            slug="celery-export-workspace",
            owner=user,
        )

        account = AvitoAccount.objects.create(workspace=workspace, name="Celery Export Account")

        create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Celery Address"],
            title="Celery title",
            description="Celery description",
            image_urls=["https://example.com/celery.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        account.refresh_from_db()
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.DIRTY)

        exported_count = export_dirty_avito_accounts_csv_task()

        account.refresh_from_db()
        self.assertEqual(exported_count, 1)
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.CLEAN)
        self.assertTrue(account.export_file_path)
        self.assertIsNotNone(account.last_exported_at)

    def test_connect_avito_account_from_token_sets_external_account_id(self):
        user = User.objects.create_user(email="avito-api-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Avito API workspace",
            slug="avito-api-workspace",
            owner=user,
        )

        account = AvitoAccount.objects.create(workspace=workspace, name="Local Avito Account")

        token = AvitoOAuthToken.objects.create(
            workspace=workspace,
            avito_account=account,
            access_token="test-access-token",
            refresh_token="test-refresh-token",
            scope="user:read items:info stats:read autoload:reports",
        )

        class FakeResponse:
            status_code = 200
            text = '{"id": 94235311}'

            def json(self):
                return {
                    "id": 94235311,
                    "name": "Петр",
                    "email": "owner@example.com",
                    "phone": "71112223344",
                    "phones": ["71112223344"],
                    "profile_url": "https://avito.ru/user/test/profile",
                }

        class FakeSession:
            def request(self, method, url, **kwargs):
                self.method = method
                self.url = url
                self.kwargs = kwargs
                return FakeResponse()

        session = FakeSession()

        connected_account = connect_avito_account_from_token(
            avito_account=account,
            token=token,
            session=session,
        )

        token.refresh_from_db()

        self.assertEqual(connected_account.external_account_id, "94235311")
        self.assertEqual(token.user_info["id"], 94235311)
        self.assertEqual(token.user_info["name"], "Петр")
        self.assertEqual(session.method, "GET")
        self.assertEqual(session.url, "https://api.avito.ru/core/v1/accounts/self")
        self.assertEqual(
            session.kwargs["headers"]["Authorization"],
            "Bearer test-access-token",
        )

    def test_import_avito_listings_for_account_creates_listings_and_is_idempotent(self):
        user = User.objects.create_user(email="avito-import-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Avito import workspace",
            slug="avito-import-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="Import Account",
            external_account_id="94235311",
        )
        AvitoOAuthToken.objects.create(
            workspace=workspace,
            avito_account=account,
            access_token="import-access-token",
            refresh_token="import-refresh-token",
            scope="items:info",
        )

        class FakeResponse:
            status_code = 200
            text = "json"

            def json(self):
                return {
                    "meta": {"page": 1, "per_page": 25},
                    "resources": [
                        {
                            "id": 24122231,
                            "title": "Кирпич облицовочный",
                            "status": "active",
                            "url": "https://www.avito.ru/item/24122231",
                            "price": 100,
                            "address": "Москва, Лесная 7",
                            "category": {"id": 19, "name": "Стройматериалы"},
                        },
                        {
                            "id": 24122232,
                            "title": "Кирпич рядовой",
                            "status": "removed",
                            "url": None,
                            "price": None,
                            "address": "Москва, Тестовая 1",
                            "category": {"id": 19, "name": "Стройматериалы"},
                        },
                    ],
                }

        class FakeSession:
            def __init__(self):
                self.calls = []

            def request(self, method, url, **kwargs):
                self.calls.append((method, url, kwargs))
                return FakeResponse()

        session = FakeSession()

        result = import_avito_listings_for_account(account, session=session)

        self.assertEqual(result.total_received, 2)
        self.assertEqual(result.created_listings, 2)
        self.assertEqual(result.updated_listings, 0)

        self.assertEqual(AvitoListing.objects.filter(avito_account=account).count(), 2)
        self.assertEqual(AdCreative.objects.filter(workspace=workspace).count(), 0)
        self.assertEqual(AdPublication.objects.filter(workspace=workspace).count(), 0)

        listing = AvitoListing.objects.get(avito_id="24122231")
        self.assertEqual(listing.status, "active")
        self.assertEqual(listing.title, "Кирпич облицовочный")
        self.assertIsNone(listing.publication)
        self.assertEqual(listing.imported_payload["price"], 100)
        self.assertEqual(listing.imported_payload["category"]["name"], "Стройматериалы")

        second_result = import_avito_listings_for_account(account, session=session)

        self.assertEqual(second_result.created_listings, 0)
        self.assertEqual(second_result.updated_listings, 2)
        self.assertEqual(AvitoListing.objects.filter(avito_account=account).count(), 2)
        self.assertEqual(AdCreative.objects.filter(workspace=workspace).count(), 0)
        self.assertEqual(AdPublication.objects.filter(workspace=workspace).count(), 0)

    def test_import_avito_account_listings_task_imports_account_listings(self):
        user = User.objects.create_user(email="avito-import-task-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Avito import task workspace",
            slug="avito-import-task-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="Import Task Account",
            external_account_id="94235311",
        )
        AvitoOAuthToken.objects.create(
            workspace=workspace,
            avito_account=account,
            access_token="import-task-access-token",
            refresh_token="import-task-refresh-token",
            scope="items:info",
        )

        class FakeResponse:
            status_code = 200
            text = "json"

            def json(self):
                return {
                    "resources": [
                        {
                            "id": 24122233,
                            "title": "Импорт через task",
                            "status": "active",
                            "url": "https://www.avito.ru/item/24122233",
                            "price": 150,
                            "address": "Москва, Task 1",
                            "category": {"id": 19, "name": "Стройматериалы"},
                        },
                    ],
                }

        class FakeSession:
            def request(self, method, url, **kwargs):
                return FakeResponse()

        result = import_avito_account_listings_task(
            account.id,
            session=FakeSession(),
        )

        self.assertEqual(result["total_received"], 1)
        self.assertEqual(result["created_listings"], 1)
        self.assertEqual(AvitoListing.objects.filter(avito_account=account).count(), 1)

    def test_link_publications_to_avito_ids_for_account_links_row_id_to_listing(self):
        user = User.objects.create_user(email="autoload-link-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Autoload link workspace",
            slug="autoload-link-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="Autoload Account",
            external_account_id="94235311",
        )
        AvitoOAuthToken.objects.create(
            workspace=workspace,
            avito_account=account,
            access_token="autoload-access-token",
            refresh_token="autoload-refresh-token",
            scope="autoload:reports",
        )

        mass_posting = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Autoload Address 1", "Autoload Address 2"],
            title="Autoload title",
            description="Autoload description",
            image_urls=["https://example.com/autoload.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        first_publication = mass_posting.publications[0]
        second_publication = mass_posting.publications[1]

        class FakeResponse:
            status_code = 200
            text = "json"

            def json(self):
                return {
                    "items": [
                        {
                            "ad_id": first_publication.row_id,
                            "avito_id": 24122241,
                        },
                        {
                            "ad_id": second_publication.row_id,
                            "avito_id": None,
                        },
                    ],
                }

        class FakeSession:
            def __init__(self):
                self.calls = []

            def request(self, method, url, **kwargs):
                self.calls.append((method, url, kwargs))
                return FakeResponse()

        session = FakeSession()

        result = link_publications_to_avito_ids_for_account(
            account,
            session=session,
        )

        first_publication.refresh_from_db()
        second_publication.refresh_from_db()

        self.assertEqual(result.total_requested, 2)
        self.assertEqual(result.linked, 1)
        self.assertEqual(result.missing, 1)
        self.assertEqual(result.conflicts, 0)
        self.assertEqual(result.created_listings, 1)
        self.assertEqual(result.updated_listings, 0)

        listing = AvitoListing.objects.get(avito_id="24122241")
        self.assertEqual(listing.publication, first_publication)
        self.assertEqual(listing.avito_account, account)
        self.assertEqual(listing.status, "published")
        self.assertEqual(first_publication.avito_listing, listing)

        self.assertFalse(hasattr(second_publication, "avito_listing"))

        self.assertEqual(session.calls[0][0], "GET")
        self.assertEqual(
            session.calls[0][1],
            "https://api.avito.ru/autoload/v2/items/avito_ids",
        )
        self.assertIn(first_publication.row_id, session.calls[0][2]["params"]["query"])
        self.assertIn(second_publication.row_id, session.calls[0][2]["params"]["query"])

        second_result = link_publications_to_avito_ids_for_account(
            account,
            session=session,
        )

        self.assertEqual(second_result.created_listings, 0)
        self.assertEqual(second_result.updated_listings, 1)
        self.assertEqual(AvitoListing.objects.filter(avito_account=account).count(), 1)

    def test_link_avito_account_publications_task_links_publications(self):
        user = User.objects.create_user(email="autoload-link-task-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Autoload link task workspace",
            slug="autoload-link-task-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="Autoload Link Task Account",
            external_account_id="94235311",
        )
        AvitoOAuthToken.objects.create(
            workspace=workspace,
            avito_account=account,
            access_token="autoload-link-task-token",
            refresh_token="autoload-link-task-refresh",
            scope="autoload:reports",
        )

        mass_posting = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Task Link Address"],
            title="Task link title",
            description="Task link description",
            image_urls=["https://example.com/task-link.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        publication = mass_posting.publications[0]

        class FakeResponse:
            status_code = 200
            text = "json"

            def json(self):
                return {
                    "items": [
                        {
                            "ad_id": publication.row_id,
                            "avito_id": 24122251,
                        },
                    ],
                }

        class FakeSession:
            def request(self, method, url, **kwargs):
                return FakeResponse()

        result = link_avito_account_publications_task(
            account.id,
            session=FakeSession(),
        )

        publication.refresh_from_db()

        self.assertEqual(result["total_requested"], 1)
        self.assertEqual(result["linked"], 1)
        self.assertEqual(result["created_listings"], 1)
        self.assertEqual(publication.avito_listing.avito_id, "24122251")

    def test_import_avito_listing_daily_stats_for_account_upserts_daily_stats(self):
        user = User.objects.create_user(email="avito-stats-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Avito stats workspace",
            slug="avito-stats-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="Stats Account",
            external_account_id="94235311",
        )
        AvitoOAuthToken.objects.create(
            workspace=workspace,
            avito_account=account,
            access_token="stats-access-token",
            refresh_token="stats-refresh-token",
            scope="stats:read",
        )

        listing = AvitoListing.objects.create(
            workspace=workspace,
            avito_account=account,
            avito_id="24122261",
            status="active",
            title="Stats listing",
        )

        class FakeResponse:
            status_code = 200
            text = "json"

            def json(self):
                return {
                    "result": {
                        "items": [
                            {
                                "itemId": "24122261",
                                "stats": [
                                    {
                                        "date": "2026-05-01",
                                        "uniqViews": 10,
                                        "uniqContacts": 2,
                                        "uniqFavorites": 1,
                                    },
                                    {
                                        "date": "2026-05-02",
                                        "uniqViews": 15,
                                        "uniqContacts": 3,
                                        "uniqFavorites": 4,
                                    },
                                ],
                            },
                        ],
                    },
                }

        class FakeSession:
            def __init__(self):
                self.calls = []

            def request(self, method, url, **kwargs):
                self.calls.append((method, url, kwargs))
                return FakeResponse()

        session = FakeSession()

        result = import_avito_listing_daily_stats_for_account(
            avito_account=account,
            date_from=date(2026, 5, 1),
            date_to=date(2026, 5, 2),
            session=session,
        )

        self.assertEqual(result.total_listings, 1)
        self.assertEqual(result.total_days, 2)
        self.assertEqual(result.created_stats, 2)
        self.assertEqual(result.updated_stats, 0)

        first_day = AvitoListingDailyStats.objects.get(
            listing=listing,
            date=date(2026, 5, 1),
        )
        self.assertEqual(first_day.views, 10)
        self.assertEqual(first_day.contacts, 2)
        self.assertEqual(first_day.favorites, 1)
        self.assertEqual(first_day.raw_metrics["uniqViews"], 10)

        self.assertEqual(session.calls[0][0], "POST")
        self.assertEqual(
            session.calls[0][1],
            "https://api.avito.ru/stats/v1/accounts/94235311/items",
        )
        self.assertEqual(session.calls[0][2]["json"]["itemIds"], [24122261])
        self.assertEqual(session.calls[0][2]["json"]["dateFrom"], "2026-05-01")
        self.assertEqual(session.calls[0][2]["json"]["dateTo"], "2026-05-02")

        second_result = import_avito_listing_daily_stats_for_account(
            avito_account=account,
            date_from=date(2026, 5, 1),
            date_to=date(2026, 5, 2),
            session=session,
        )

        self.assertEqual(second_result.created_stats, 0)
        self.assertEqual(second_result.updated_stats, 2)
        self.assertEqual(AvitoListingDailyStats.objects.filter(listing=listing).count(), 2)

    def test_import_avito_account_daily_stats_task_imports_stats(self):
        user = User.objects.create_user(email="avito-stats-task-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Avito stats task workspace",
            slug="avito-stats-task-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(
            workspace=workspace,
            name="Stats Task Account",
            external_account_id="94235311",
        )
        AvitoOAuthToken.objects.create(
            workspace=workspace,
            avito_account=account,
            access_token="stats-task-access-token",
            refresh_token="stats-task-refresh-token",
            scope="stats:read",
        )

        listing = AvitoListing.objects.create(
            workspace=workspace,
            avito_account=account,
            avito_id="24122271",
            status="active",
            title="Stats task listing",
        )

        class FakeResponse:
            status_code = 200
            text = "json"

            def json(self):
                return {
                    "result": {
                        "items": [
                            {
                                "itemId": "24122271",
                                "stats": [
                                    {
                                        "date": "2026-05-03",
                                        "uniqViews": 20,
                                        "uniqContacts": 4,
                                        "uniqFavorites": 2,
                                    },
                                ],
                            },
                        ],
                    },
                }

        class FakeSession:
            def request(self, method, url, **kwargs):
                return FakeResponse()

        result = import_avito_account_daily_stats_task(
            account.id,
            "2026-05-03",
            "2026-05-03",
            session=FakeSession(),
        )

        stats = AvitoListingDailyStats.objects.get(
            listing=listing,
            date=date(2026, 5, 3),
        )

        self.assertEqual(result["total_listings"], 1)
        self.assertEqual(result["total_days"], 1)
        self.assertEqual(result["created_stats"], 1)
        self.assertEqual(stats.views, 20)
        self.assertEqual(stats.contacts, 4)
        self.assertEqual(stats.favorites, 2)

    def test_archive_stale_publications_archives_only_old_inactive_publications(self):
        user = User.objects.create_user(email="cleanup-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Cleanup workspace",
            slug="cleanup-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Cleanup Account")

        old_batch = AdBatch.objects.create(
            workspace=workspace,
            source=AdBatch.Source.MANUAL,
            status=AdBatch.Status.COMPLETED,
        )

        old_creative = AdCreative.objects.create(
            workspace=workspace,
            batch=old_batch,
            source=AdCreative.Source.MANUAL,
            title="Old cleanup title",
            description="Old cleanup description",
            image_urls=[],
            base_data={},
            option_data={},
        )

        old_publication = AdPublication.objects.create(
            workspace=workspace,
            avito_account=account,
            creative=old_creative,
            batch=old_batch,
            source=AdPublication.Source.MANUAL,
            status=AdPublication.Status.PAUSED,
            row_id="OLD-ROW",
            address="Old address",
        )
        fresh_publication = AdPublication.objects.create(
            workspace=workspace,
            avito_account=account,
            creative=old_creative,
            batch=old_batch,
            source=AdPublication.Source.MANUAL,
            status=AdPublication.Status.PAUSED,
            row_id="FRESH-ROW",
            address="Fresh address",
        )
        active_publication = AdPublication.objects.create(
            workspace=workspace,
            avito_account=account,
            creative=old_creative,
            batch=old_batch,
            source=AdPublication.Source.MANUAL,
            status=AdPublication.Status.ACTIVE,
            row_id="ACTIVE-ROW",
            address="Active address",
        )
        linked_active_publication = AdPublication.objects.create(
            workspace=workspace,
            avito_account=account,
            creative=old_creative,
            batch=old_batch,
            source=AdPublication.Source.MANUAL,
            status=AdPublication.Status.PAUSED,
            row_id="LINKED-ROW",
            address="Linked address",
        )

        AvitoListing.objects.create(
            workspace=workspace,
            avito_account=account,
            publication=linked_active_publication,
            avito_id="24122281",
            status="active",
            title="Linked active listing",
        )

        old_dt = timezone.now() - timedelta(days=90)
        fresh_dt = timezone.now() - timedelta(days=5)

        AdPublication.objects.filter(id=old_publication.id).update(created_at=old_dt)
        AdPublication.objects.filter(id=active_publication.id).update(created_at=old_dt)
        AdPublication.objects.filter(id=linked_active_publication.id).update(created_at=old_dt)
        AdPublication.objects.filter(id=fresh_publication.id).update(created_at=fresh_dt)

        result = archive_stale_publications(
            workspace=workspace,
            older_than_days=60,
        )

        old_publication.refresh_from_db()
        fresh_publication.refresh_from_db()
        active_publication.refresh_from_db()
        linked_active_publication.refresh_from_db()

        self.assertEqual(result.archived_publications, 1)
        self.assertEqual(old_publication.status, AdPublication.Status.ARCHIVED)
        self.assertIsNotNone(old_publication.archived_at)

        self.assertEqual(fresh_publication.status, AdPublication.Status.PAUSED)
        self.assertEqual(active_publication.status, AdPublication.Status.ACTIVE)
        self.assertEqual(linked_active_publication.status, AdPublication.Status.PAUSED)
        self.assertEqual(AvitoListing.objects.filter(publication=linked_active_publication).count(), 1)

    def test_product_random_api_uses_autogeneration_pipeline(self):
        user = User.objects.create_user(email="product-random-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Product random workspace",
            slug="product-random-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Product Random Account")
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Manual API task",
            is_active=False,
            titles=["Manual API title"],
            descriptions={"0": "Manual API description {{ TITLE }} / {{ SKU }}"},
            addresses=["Manual API Address 1", "Manual API Address 2"],
            price=100,
        )
        self.attach_task_images(task, main_urls=["https://example.com/manual-api.jpg"])
        task.avito_accounts.add(account)

        response = client.post(
            reverse("product-random-api", args=[task.id]),
            {},
            format="json",
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        task.refresh_from_db()
        account.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["publications_count"], 2)
        self.assertEqual(response.data["csv_export_status"], "queued")
        self.assertIsNotNone(response.data["run_id"])
        self.assertIsNotNone(response.data["creative_id"])
        self.assertIsNotNone(response.data["batch_id"])

        run = AdGenerationTaskRun.objects.get(id=response.data["run_id"])
        self.assertEqual(run.task, task)
        self.assertEqual(run.triggered_by, AdGenerationTaskRun.TriggeredBy.MANUAL)
        self.assertEqual(run.status, AdGenerationTaskRun.Status.SUCCESS)

        self.assertEqual(AdCreative.objects.filter(task=task).count(), 1)
        self.assertEqual(AdPublication.objects.filter(task=task).count(), 2)
        self.assertEqual(account.export_status, AvitoAccount.ExportStatus.DIRTY)

        self.assertEqual(task.last_run_status, AdGenerationTask.LastRunStatus.SUCCESS)
        self.assertIsNotNone(task.last_run_at)
        self.assertEqual(task.last_successful_run_at, task.last_run_at)

    def test_product_random_api_returns_400_for_invalid_task(self):
        user = User.objects.create_user(email="product-random-invalid-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Product random invalid workspace",
            slug="product-random-invalid-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Invalid manual API task",
            is_active=False,
            titles=[],
            descriptions={"0": "Description without title"},
            addresses=["Invalid Address"],
            price=100,
        )
        self.attach_task_images(task, main_urls=["https://example.com/invalid.jpg"])

        response = client.post(
            reverse("product-random-api", args=[task.id]),
            {},
            format="json",
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        task.refresh_from_db()

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["error"], "У задачи нет заголовков")
        self.assertEqual(task.last_run_status, AdGenerationTask.LastRunStatus.ERROR)
        self.assertEqual(task.last_run_error, "У задачи нет заголовков")

    def test_toggle_product_active_activates_task_and_recalculates_next_update_time(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="toggle-activate-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Toggle activate workspace",
            slug="toggle-activate-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Inactive task",
            is_active=False,
            titles=["Toggle title"],
            descriptions={"0": "Toggle description"},
            addresses=["Toggle Address"],
            price=100,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=None,
        )
        self.attach_task_images(task, main_urls=["https://example.com/toggle.jpg"])

        with patch(
                "avitotask.services.ad_schedule.timezone.now",
                return_value=datetime(2026, 5, 4, 9, 59, tzinfo=tz),
        ):
            response = client.post(
                reverse("toggle-product-active-api", args=[task.id]),
                {"action": "activate"},
                format="json",
                HTTP_X_WORKSPACE_ID=str(workspace.id),
            )

        task.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["status"], "ok")
        self.assertEqual(response.data["active"], True)
        self.assertEqual(response.data["next_update_time"], "2026-05-04T10:00:00+03:00")

        self.assertTrue(task.is_active)
        self.assertEqual(task.next_update_time, datetime(2026, 5, 4, 10, 0, tzinfo=tz))

    def test_toggle_product_active_deactivates_task_and_clears_next_update_time(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="toggle-deactivate-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Toggle deactivate workspace",
            slug="toggle-deactivate-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Active task",
            is_active=True,
            schedule={
                "frequency": 1,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            next_update_time=datetime(2026, 5, 4, 10, 0, tzinfo=tz),
        )

        response = client.post(
            reverse("toggle-product-active-api", args=[task.id]),
            {"action": "deactivate"},
            format="json",
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        task.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["status"], "ok")
        self.assertEqual(response.data["active"], False)
        self.assertIsNone(response.data["next_update_time"])

        self.assertFalse(task.is_active)
        self.assertIsNone(task.next_update_time)

    def test_toggle_product_active_returns_400_for_invalid_schedule(self):
        user = User.objects.create_user(email="toggle-invalid-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Toggle invalid workspace",
            slug="toggle-invalid-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Invalid schedule task",
            is_active=False,
            schedule={
                "frequency": 1,
                "days": [None, None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
        )

        response = client.post(
            reverse("toggle-product-active-api", args=[task.id]),
            {"action": "activate"},
            format="json",
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        task.refresh_from_db()

        self.assertEqual(response.status_code, 400)
        self.assertIn("error", response.data)
        self.assertFalse(task.is_active)
        self.assertIsNone(task.next_update_time)

    def test_products_api_uses_ad_generation_tasks_and_returns_run_fields(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="products-api-task-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Products API task workspace",
            slug="products-api-task-workspace",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="New generation task",
            is_active=True,
            url="https://example.com/new-task",
            titles=["API title"],
            descriptions={"0": "API description"},
            addresses=["API Address"],
            price=100,
            schedule={
                "frequency": 2,
                "days": ["10:00", None, None, None, None, None, None],
            },
            schedule_anchor_date=date(2026, 5, 4),
            schedule_timezone="Europe/Moscow",
            publication_interval_days=14,
            next_update_time=datetime(2026, 5, 18, 10, 0, tzinfo=tz),
            last_run_at=datetime(2026, 5, 4, 10, 0, tzinfo=tz),
            last_successful_run_at=datetime(2026, 5, 4, 10, 0, tzinfo=tz),
            last_run_status=AdGenerationTask.LastRunStatus.SUCCESS,
            last_run_error=None,
        )
        main_assets, additional_assets = self.attach_task_images(
            task,
            main_urls=["https://example.com/api.jpg"],
            additional_urls=["https://example.com/api-extra.jpg"],
        )

        response = client.get(
            reverse("product-api-list"),
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["id"], task.id)
        self.assertEqual(response.data[0]["name"], "New generation task")
        self.assertEqual(response.data[0]["activate"], True)
        self.assertEqual(response.data[0]["schedule"], {
            "frequency": 2,
            "days": ["10:00", None, None, None, None, None, None],
        })
        self.assertEqual(response.data[0]["next_update_time"], "2026-05-18T10:00:00+03:00")
        self.assertEqual(response.data[0]["last_run_at"], "2026-05-04T10:00:00+03:00")
        self.assertEqual(response.data[0]["last_successful_run_at"], "2026-05-04T10:00:00+03:00")
        self.assertEqual(response.data[0]["last_run_status"], "success")
        self.assertIsNone(response.data[0]["last_run_error"])
        self.assertEqual(response.data[0]["main_images"], ["https://example.com/api.jpg"])
        self.assertEqual(response.data[0]["additional_images"], ["https://example.com/api-extra.jpg"])
        self.assertEqual(response.data[0]["main_image_asset_ids"], [main_assets[0].id])
        self.assertEqual(response.data[0]["additional_image_asset_ids"], [additional_assets[0].id])

    def test_products_api_create_writes_ad_generation_task_not_legacy_product(self):
        tz = ZoneInfo("Europe/Moscow")

        user = User.objects.create_user(email="products-api-create-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Products API create workspace",
            slug="products-api-create-workspace",
            owner=user,
        )
        account = AvitoAccount.objects.create(workspace=workspace, name="Products API Account")
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)
        main_asset = self.create_image_asset(
            workspace=workspace,
            url="https://example.com/created.jpg",
            uploaded_by=user,
        )

        with patch(
                "avitotask.services.ad_schedule.timezone.now",
                return_value=datetime(2026, 5, 4, 9, 0, tzinfo=tz),
        ):
            response = client.post(
                reverse("product-api-list"),
                {
                    "name": "Created generation task",
                    "url": "https://example.com/created-task",
                    "activate": True,
                    "price": 1000,
                    "titles": ["Created title"],
                    "descriptions": ["Created description"],
                    "main_image_asset_ids": [main_asset.id],
                    "additional_image_asset_ids": [],
                    "addresses": ["Created Address"],
                    "schedule": {
                        "frequency": 3,
                        "days": ["10:00", None, None, None, None, None, None],
                    },
                    "schedule_anchor_date": "2026-05-04",
                    "schedule_timezone": "Europe/Moscow",
                    "avito_account_ids": [account.id],
                },
                format="json",
                HTTP_X_WORKSPACE_ID=str(workspace.id),
            )

        self.assertEqual(response.status_code, 201)

        task = AdGenerationTask.objects.get(workspace=workspace)
        self.assertEqual(task.name, "Created generation task")
        self.assertEqual(task.is_active, True)
        self.assertEqual(task.publication_interval_days, 21)
        self.assertEqual(task.schedule, {
            "frequency": 3,
            "days": ["10:00", None, None, None, None, None, None],
        })
        self.assertEqual(task.schedule_anchor_date, date(2026, 5, 4))
        self.assertEqual(task.schedule_timezone, "Europe/Moscow")
        self.assertEqual(task.next_update_time, datetime(2026, 5, 4, 10, 0, tzinfo=tz))
        self.assertEqual(list(task.avito_accounts.values_list("id", flat=True)), [account.id])
        self.assertEqual(list(task.main_image_assets.values_list("id", flat=True)), [main_asset.id])

        self.assertEqual(response.data["activate"], True)
        self.assertEqual(response.data["last_run_status"], "idle")
        self.assertEqual(response.data["next_update_time"], "2026-05-04T10:00:00+03:00")

    def test_legacy_schedule_price_updates_is_not_in_beat_schedule(self):
        beat_schedule = celery_app.conf.beat_schedule

        self.assertNotIn("schedule_price_updates_every_minute", beat_schedule)
        self.assertIn("run_due_ad_generation_tasks_every_minute", beat_schedule)
        self.assertEqual(
            beat_schedule["run_due_ad_generation_tasks_every_minute"]["task"],
            "avitotask.tasks.run_due_ad_generation_tasks",
        )

    def test_creatives_queryset_does_not_return_import_creatives(self):
        user = User.objects.create_user(email="creative-source-owner@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Creative source workspace",
            slug="creative-source-workspace",
            owner=user,
        )

        import_batch = AdBatch.objects.create(
            workspace=workspace,
            source=AdBatch.Source.IMPORT,
            status=AdBatch.Status.COMPLETED,
        )

        AdCreative.objects.create(
            workspace=workspace,
            batch=import_batch,
            source="import",
            title="Imported Avito creative",
            description="Should not be visible",
            image_urls=[],
            base_data={},
            option_data={},
        )

        manual_result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[AvitoAccount.objects.create(workspace=workspace, name="Account")],
            addresses=["Address"],
            title="Manual creative",
            description="Manual description",
            image_urls=["https://example.com/manual.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        queryset = AdCreative.objects.filter(workspace=workspace).exclude(source="import")

        self.assertEqual(queryset.count(), 1)
        self.assertEqual(queryset.first(), manual_result.creative)

    def test_manual_mass_posting_rejects_recent_creative_when_two_of_three_fields_match(self):
        user = User.objects.create_user(email="dedupe-manual-owner@example.com", password="test")
        workspace = Workspace.objects.create(name="Dedupe manual", slug="dedupe-manual", owner=user)
        account = AvitoAccount.objects.create(workspace=workspace, name="Account")

        create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Address 1"],
            title="Same title",
            description="Same description",
            image_urls=["https://example.com/first.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        with self.assertRaisesMessage(AdGenerationError, "Похожий креатив уже существует"):
            create_manual_mass_posting(
                workspace=workspace,
                user=user,
                avito_accounts=[account],
                addresses=["Address 2"],
                title="Same title",
                description="Same description",
                image_urls=["https://example.com/second.jpg"],
                base_data={"Price": 700},
                option_data={},
            )

        self.assertEqual(AdCreative.objects.filter(workspace=workspace).count(), 1)
        self.assertEqual(AdPublication.objects.filter(workspace=workspace).count(), 1)

    def test_manual_mass_posting_allows_similar_creative_after_lookback_period(self):
        user = User.objects.create_user(email="dedupe-old-owner@example.com", password="test")
        workspace = Workspace.objects.create(name="Dedupe old", slug="dedupe-old", owner=user)
        account = AvitoAccount.objects.create(workspace=workspace, name="Account")

        first_result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Address 1"],
            title="Same title",
            description="Same description",
            image_urls=["https://example.com/first.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        AdCreative.objects.filter(id=first_result.creative.id).update(
            created_at=timezone.now() - timedelta(days=31)
        )

        second_result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Address 2"],
            title="Same title",
            description="Same description",
            image_urls=["https://example.com/second.jpg"],
            base_data={"Price": 700},
            option_data={},
        )

        self.assertNotEqual(first_result.creative.id, second_result.creative.id)
        self.assertEqual(AdCreative.objects.filter(workspace=workspace).count(), 2)

    def test_generate_ads_from_task_retries_when_recent_creative_matches_two_of_three_fields(self):
        user = User.objects.create_user(email="dedupe-auto-owner@example.com", password="test")
        workspace = Workspace.objects.create(name="Dedupe auto", slug="dedupe-auto", owner=user)
        account = AvitoAccount.objects.create(workspace=workspace, name="Account")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Dedupe task",
            is_active=True,
            titles=["Same title", "Fresh title"],
            descriptions={
                "1": "Same description {{ TITLE }} / {{ SKU }}",
                "2": "Fresh description {{ TITLE }} / {{ SKU }}",
            },
            addresses=["Address 1"],
            price=100,
        )
        self.attach_task_images(task, main_urls=["https://example.com/same.jpg"])
        task.avito_accounts.add(account)

        existing_dedupe = build_creative_dedupe_data(
            title="Same title",
            description="Same description Same title / __SKU__",
            image_urls=["https://example.com/same.jpg"],
        )

        AdCreative.objects.create(
            workspace=workspace,
            task=task,
            source=AdCreative.Source.AUTO,
            title="Same title",
            description="Same description Same title / OLD-SKU",
            image_urls=["https://example.com/same.jpg"],
            base_data={"Price": 100},
            option_data={},
            identity_hash="old",
            dedupe_title=existing_dedupe["dedupe_title"],
            dedupe_description=existing_dedupe["dedupe_description"],
            dedupe_images_hash=existing_dedupe["dedupe_images_hash"],
        )

        result = generate_ads_from_task(task.id, workspace=workspace, user=user)

        self.assertNotEqual(result.creative.title, "Same title")
        self.assertEqual(result.creative.title, "Fresh title")
        self.assertEqual(AdCreative.objects.filter(workspace=workspace).count(), 2)
        self.assertEqual(AdPublication.objects.filter(workspace=workspace).count(), 1)

    def test_generate_ads_from_task_retries_when_recent_creative_matches_two_of_three_fields(self):
        user = User.objects.create_user(email="dedupe-auto-owner@example.com", password="test")
        workspace = Workspace.objects.create(name="Dedupe auto", slug="dedupe-auto", owner=user)
        account = AvitoAccount.objects.create(workspace=workspace, name="Account")

        task = AdGenerationTask.objects.create(
            workspace=workspace,
            name="Dedupe task",
            is_active=True,
            titles=["Same title", "Fresh title"],
            descriptions={
                "1": "Same description {{ TITLE }} / {{ SKU }}",
                "2": "Fresh description {{ TITLE }} / {{ SKU }}",
            },
            addresses=["Address 1"],
            price=100,
        )
        self.attach_task_images(
            task,
            main_urls=[
                "https://example.com/same.jpg",
                "https://example.com/fresh.jpg",
            ],
        )
        task.avito_accounts.add(account)

        existing_dedupe = build_creative_dedupe_data(
            title="Same title",
            description="Same description Same title / __SKU__",
            image_urls=["https://example.com/same.jpg"],
        )

        AdCreative.objects.create(
            workspace=workspace,
            task=task,
            source=AdCreative.Source.AUTO,
            title="Same title",
            description="Same description Same title / OLD-SKU",
            image_urls=["https://example.com/same.jpg"],
            base_data={"Price": 100},
            option_data={},
            identity_hash="old",
            dedupe_title=existing_dedupe["dedupe_title"],
            dedupe_description=existing_dedupe["dedupe_description"],
            dedupe_images_hash=existing_dedupe["dedupe_images_hash"],
        )

        choices = iter([
            "Same title",
            "Same description {{ TITLE }} / {{ SKU }}",
            "https://example.com/same.jpg",
            "Fresh title",
            "Fresh description {{ TITLE }} / {{ SKU }}",
            "https://example.com/fresh.jpg",
        ])

        def choose_deterministic_value(values):
            selected = next(choices)

            if values and hasattr(values[0], "url"):
                return next(asset for asset in values if asset.url == selected)

            return selected

        with patch("avitotask.services.ad_generation.random.choice", side_effect=choose_deterministic_value):
            result = generate_ads_from_task(task.id, workspace=workspace, user=user)

        self.assertEqual(result.creative.title, "Fresh title")
        self.assertEqual(result.creative.image_urls, ["https://example.com/fresh.jpg"])
        self.assertEqual(AdCreative.objects.filter(workspace=workspace).count(), 2)
        self.assertEqual(AdPublication.objects.filter(workspace=workspace).count(), 1)

    def test_ad_creatives_api_returns_empty_for_import_source_filter(self):
        user = User.objects.create_user(email="creative-api-import@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Creative API import filter",
            slug="creative-api-import-filter",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        AdCreative.objects.create(
            workspace=workspace,
            source="import",
            title="Imported creative",
            description="Should not be returned",
            image_urls=[],
            base_data={},
            option_data={},
        )

        response = client.get(
            reverse("ad-creative-api-list"),
            {"source": "import"},
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 0)
        self.assertEqual(response.data["results"], [])

    def test_ad_publications_api_returns_empty_for_import_source_filter(self):
        user = User.objects.create_user(email="publication-api-import@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Publication API import filter",
            slug="publication-api-import-filter",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        account = AvitoAccount.objects.create(workspace=workspace, name="Account")

        import_creative = AdCreative.objects.create(
            workspace=workspace,
            source="import",
            title="Imported creative",
            description="Should not be returned",
            image_urls=[],
            base_data={},
            option_data={},
        )

        AdPublication.objects.create(
            workspace=workspace,
            avito_account=account,
            creative=import_creative,
            source="import",
            status=AdPublication.Status.DRAFT,
            address="Imported address",
            overrides={},
        )

        response = client.get(
            reverse("ad-publication-api-list"),
            {"source": "import"},
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 0)
        self.assertEqual(response.data["results"], [])

    def test_ad_publications_api_searches_by_creative_title_and_avito_id(self):
        user = User.objects.create_user(email="publication-api-search@example.com", password="test")
        workspace = Workspace.objects.create(
            name="Publication API search",
            slug="publication-api-search",
            owner=user,
        )
        client = self.create_api_client_for_workspace(user=user, workspace=workspace)

        account = AvitoAccount.objects.create(workspace=workspace, name="Account")

        result = create_manual_mass_posting(
            workspace=workspace,
            user=user,
            avito_accounts=[account],
            addresses=["Search Address"],
            title="Searchable Creative Title",
            description="Description",
            image_urls=["https://example.com/search.jpg"],
            base_data={"Price": 500},
            option_data={},
        )

        AvitoListing.objects.create(
            workspace=workspace,
            avito_account=account,
            publication=result.publications[0],
            avito_id="777888999",
            status="published",
            title="Avito listing",
        )

        title_response = client.get(
            reverse("ad-publication-api-list"),
            {"search": "Searchable Creative"},
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )
        avito_response = client.get(
            reverse("ad-publication-api-list"),
            {"search": "777888999"},
            HTTP_X_WORKSPACE_ID=str(workspace.id),
        )

        self.assertEqual(title_response.status_code, 200)
        self.assertEqual(title_response.data["count"], 1)
        self.assertEqual(title_response.data["results"][0]["creative_title"], "Searchable Creative Title")

        self.assertEqual(avito_response.status_code, 200)
        self.assertEqual(avito_response.data["count"], 1)
        self.assertEqual(avito_response.data["results"][0]["avito_id"], "777888999")
