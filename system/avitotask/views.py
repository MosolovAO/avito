import json
import os
from datetime import datetime, timedelta

from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404

from .models import Product1, Product, ProductOptions, ProductOptionAssignment, Project
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


def product_random(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    product_id = product.id
    product_address = product.addresses
    products1 = Product1.objects.filter(task_id=product_id)

    projects = product.projects.all()

    # Получаем записи за последние 31 день с указанным task_id
    cutoff_date = datetime.now().date() - timedelta(days=31)
    existing_records = Product1.objects.filter(
        Q(task_id=product_id) & Q(created_date__gte=cutoff_date)
    ).count()

    if existing_records >= product.possible_combinations:
        return JsonResponse({'error': 'Все возможные комбинации уже созданы.'}, status=400,
                            json_dumps_params={'ensure_ascii': False})

    while True:
        # Генерация случайных данных
        product_title = product.random_title()
        product_descriptions = product.random_description()
        product_project = product.projects
        product_images = [
            product.random_main_image(),
            *product.random_additional_image(),
        ]

        matches = 0

        # Проверка совпадений
        for product1 in products1:
            matches = 0

            # Проверка совпадения по заголовку
            if product1.title == product_title:
                matches += 1

            # Проверка URL с использованием пересечения множеств
            unique_elements = set(product1.urls).intersection(set(product_images))
            if len(unique_elements) <= 5:
                matches += 1

            # Проверка совпадения по описанию
            if product1.description == product_descriptions:
                matches += 1

            # Если совпадений >= 2, пропускаем генерацию
            if matches >= 2:
                break

        if matches < 2:
            # Если условие выполнено, создаем новый продукт
            new_product = Product1.objects.create(
                title=product_title,
                urls=product_images,
                description=product_descriptions,
                task_id=product_id
            )
            new_product.save()

            # Создание или обновление Excel файла

            for project in projects:
                file_path = f"{project}_avito_autoload.xlsx"

                for product_addres in product_address:

                    if os.path.exists(file_path):
                        wb = load_workbook(file_path)
                        ws = wb.active
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Products"
                        # Добавляем заголовки, если файл создается впервые
                        ws.append(
                            ["Title", "ImageUrls", "Description", "Category", "Price", "ListingFee", "EMail",
                             "ContactPhone", "ManagerName", "AvitoStatus", "CompanyName", "ContactMethod", "AdType",
                             "Availability", "Address"])

                    # Добавляем данные нового продукта
                    row_data = {
                        "Title": product_title,
                        "ImageUrls": " | ".join(product_images),
                        "Description": product_descriptions,
                        "Category": product.category,
                        "Price": product.price,
                        "ListingFee": product.listingfee,
                        "EMail": product.email,
                        "ContactPhone": int(product.contactphone),
                        "ManagerName": product.managername,
                        "AvitoStatus": product.avitostatus,
                        "CompanyName": product.companyname,
                        "ContactMethod": product.contactmethod,
                        "AdType": product.adtype,
                        "Availability": product.availability,
                        "Address": product_addres,
                    }

                    # Обработка опций через ProductOptionAssignment
                    assignments = product.productoptionassignment_set.select_related('option').all()
                    product_options = {
                        assignment.option.option_title: assignment.selected_value
                        for assignment in assignments
                    }
                    row_data.update(product_options)

                    # Проверяем существующие заголовки столбцов
                    headers = [cell.value for cell in ws[1]]
                    for option_title in product_options.keys():
                        if option_title not in headers:
                            ws.cell(row=1, column=len(headers) + 1, value=option_title)
                            headers.append(option_title)

                    # Добавляем строку с данными
                    row = [row_data.get(header, "") for header in headers]
                    ws.append(row)

                    # Сохраняем файл
                    wb.save(file_path)

            return JsonResponse({
                'matches': matches,
                'random_title': product_title,
                'product_images': product_images,
                'product_descriptions': product_descriptions,
                'product_category': product.category,
            }, json_dumps_params={'ensure_ascii': False})


def product_add(request):
    """
    Рендерит многоэтапную форму для добавления или редактирования продукта.
    """
    # Передаём продукт и флаг редактирования в шаблон

    projects = Project.objects.all()

    context = {'projects': projects}

    return render(request, 'add_product.html', context=context)


def product_edit(request, product_id):
    """
    Рендерит многоэтапную форму для редактирования продукта.
    """
    product = get_object_or_404(Product, id=product_id)
    product.addresses = "\n".join([address.strip() for address in product.addresses])

    return render(request, 'edit_product.html', {
        'product_id': product.id,
        'product': product,
    })


def product_list(request):
    """
    Рендерит список всех задач.
    """
    products = Product.objects.prefetch_related('projects').all()

    return render(request, 'product_list.html', {'products': products})


def get_product_options(request):
    """
    Возвращает список доступных опций и их значений.
    """
    options = ProductOptions.objects.all().values('id', 'option_title', 'option_value')
    return JsonResponse(list(options), safe=False)


@csrf_exempt
def finalize_product_form(request):
    """
    Обрабатывает сохранение продукта (создание или обновление).
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('product_data'))
            product_id = data.get('product_id')  # Получаем ID продукта, если передан

            if product_id:  # Обновляем существующий продукт
                product = get_object_or_404(Product, id=product_id)
                product.titles = data['titles']
                product.main_images = data['main_images']
                product.additional_images = data['additional_images']
                product.descriptions = data['descriptions']
                product.addresses = data['addresses']
                product.category = data['category']
                product.listingfee = data['listingfee']
                product.email = data['email']
                product.contactphone = data['contactphone']
                product.managername = data['managername']
                product.avitostatus = data['avitostatus']
                product.companyname = data['companyname']
                product.contactmethod = data['contactmethod']
                product.adtype = data['adtype']
                product.availability = data['availability']
                product.price = data['price']
                product.price_max = data['price_max']
                product.price_min = data['price_min']
                product.price_step = data['price_step']

                product.save()

                # Обновляем опции
                ProductOptionAssignment.objects.filter(product=product).delete()
                for option_data in data.get('options', []):
                    option = ProductOptions.objects.get(id=option_data['option_id'])
                    ProductOptionAssignment.objects.create(
                        product=product,
                        option=option,
                        selected_value=option_data['value']
                    )
            else:  # Создаем новый продукт
                selected_projects = data['projects']
                projects = Project.objects.filter(id__in=selected_projects)

                product = Product.objects.create(
                    name=f"Задача {random.randint(0, 20000000000)}",
                    titles=data['titles'],
                    main_images=data['main_images'],
                    additional_images=data['additional_images'],
                    descriptions=data['descriptions'],
                    addresses=data['addresses'],
                    category=data['category'],
                    listingfee=data['listingfee'],
                    email=data['email'],
                    contactphone=data['contactphone'],
                    managername=data['managername'],
                    avitostatus=data['avitostatus'],
                    companyname=data['companyname'],
                    contactmethod=data['contactmethod'],
                    adtype=data['adtype'],
                    availability=data['availability'],
                    price=data['price'],
                    price_max=data['price_max'],
                    price_min=data['price_min'],
                    price_step=data['price_step'],
                    possible_combinations=data['possible_combinations'],

                )

                product.projects.set(projects)

                for option_data in data.get('options', []):
                    try:
                        option = ProductOptions.objects.get(id=option_data['option_id'])
                        ProductOptionAssignment.objects.create(
                            product=product,
                            option=option,
                            selected_value=option_data['value']
                        )
                    except ProductOptions.DoesNotExist:
                        print(f"Опция с ID {option_data['option_id']} не найдена.")

            return JsonResponse({'message': 'Продукт успешно сохранён!'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def get_product_data(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    data = {
        'titles': product.titles,
        'main_images': product.main_images,
        'additional_images': product.additional_images,
        'descriptions': product.descriptions,
        'addresses': product.addresses,
        'options': [
            {
                'option_title': assignment.option.option_title,
                'selected_value': assignment.selected_value,
            }
            for assignment in product.productoptionassignment_set.select_related('option')
        ],
    }
    return JsonResponse(data)
